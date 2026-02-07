import os
import re
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO

import requests
import json
import calendar
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText


# =========================
# Settings / Secrets
# =========================
EXCEL_URL = os.environ.get("EXCEL_URL", "").strip()

SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "").strip()
SMTP_PASS = os.environ.get("SMTP_PASS", "").strip()
MAIL_FROM = os.environ.get("MAIL_FROM", "").strip()
MAIL_TO = os.environ.get("MAIL_TO", "").strip()

SUBSCRIBE_URL = os.environ.get("SUBSCRIBE_URL", "").strip()
SUBSCRIBE_TOKEN = os.environ.get("SUBSCRIBE_TOKEN", "").strip()

PAGES_BASE_URL = os.environ.get("PAGES_BASE_URL", "").strip()  # optional
TZ = ZoneInfo("Asia/Muscat")
AUTO_OPEN_ACTIVE_SHIFT_IN_FULL = True
# Excel sheets
DEPARTMENTS = [
    ("Officers", "Officers"),
    ("Supervisors", "Supervisors"),
    ("Load Control", "Load Control"),
    ("Export Checker", "Export Checker"),
    ("Export Operators", "Export Operators"),
]

# For day-row matching only
DAYS = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

SHIFT_MAP = {
    "MN06": ("üåÖ Morning (MN06)", "ÿµÿ®ÿßÿ≠"),
    "ME06": ("üåÖ Morning (ME06)", "ÿµÿ®ÿßÿ≠"),
    "ME07": ("üåÖ Morning (ME07)", "ÿµÿ®ÿßÿ≠"),
    "MN12": ("üåÜ Afternoon (MN12)", "ÿ∏Ÿáÿ±"),
    "AN13": ("üåÜ Afternoon (AN13)", "ÿ∏Ÿáÿ±"),
    "AE14": ("üåÜ Afternoon (AE14)", "ÿ∏Ÿáÿ±"),
    "NN21": ("üåô Night (NN21)", "ŸÑŸäŸÑ"),
    "NE22": ("üåô Night (NE22)", "ŸÑŸäŸÑ"),
}

GROUP_ORDER = ["ÿµÿ®ÿßÿ≠", "ÿ∏Ÿáÿ±", "ŸÑŸäŸÑ", "ŸÖŸÜÿßŸàÿ®ÿßÿ™", "ÿ±ÿßÿ≠ÿ©", "ÿ•ÿ¨ÿßÿ≤ÿßÿ™", "ÿ™ÿØÿ±Ÿäÿ®", "ÿ£ÿÆÿ±Ÿâ"]
DEPT_COLORS = ["#2563eb", "#7c3aed", "#0891b2", "#059669", "#dc2626", "#ea580c"]

# ÿ£ÿ∂ŸÅ Ÿáÿ∞ÿß ŸÅŸä ÿ®ÿØÿßŸäÿ© ÿßŸÑŸÖŸÑŸÅ ŸÖÿπ ÿßŸÑŸÖÿ™ÿ∫Ÿäÿ±ÿßÿ™ ÿßŸÑÿ´ÿßÿ®ÿ™ÿ© ÿßŸÑÿ£ÿÆÿ±Ÿâ
DEPT_EMAIL_COLORS = {
    "Emergency": "#dc2626",      # ÿ£ÿ≠ŸÖÿ±
    "ICU": "#7c3aed",           # ÿ®ŸÜŸÅÿ≥ÿ¨Ÿä
    "Surgery": "#2563eb",        # ÿ£ÿ≤ÿ±ŸÇ
    "Pediatrics": "#16a34a",     # ÿ£ÿÆÿ∂ÿ±
    "Radiology": "#ea580c",      # ÿ®ÿ±ÿ™ŸÇÿßŸÑŸä
    # ÿ£ÿ∂ŸÅ ÿ®ŸÇŸäÿ© ÿßŸÑÿ£ŸÇÿ≥ÿßŸÖ...
}



def build_pretty_email_html(active_group: str, now: datetime, rows_by_dept: list, pages_base: str) -> str:
    """
    Email-safe HTML (tables + inline styles) with department header colors like the site.
    rows_by_dept = [{"dept": str, "rows": [{"name": str, "shift": str}, ...]}, ...]
    """
    # Date label (robust for runners)
    try:
        date_label = now.strftime("%-d %B %Y")
    except Exception:
        date_label = now.strftime("%d %B %Y")

    sent_time = now.strftime("%H:%M")

    # Shift theme (for status color)
    def shift_theme(g: str):
        if g == "ÿµÿ®ÿßÿ≠":
            return ("#fef3c7", "#f59e0b55", "#92400e")
        if g == "ÿ∏Ÿáÿ±":
            return ("#ffedd5", "#f9731655", "#9a3412")
        if g == "ŸÑŸäŸÑ":
            return ("#ede9fe", "#8b5cf655", "#5b21b6")
        return ("#e0e7ff", "#6366f155", "#3730a3")

    bg, border, textc = shift_theme(active_group)

    dept_blocks = []
    total_now = 0
    depts_now = 0

    for item in rows_by_dept:
        dept = item.get("dept", "")
        rows = item.get("rows", []) or []
        if not rows:
            continue

        depts_now += 1
        total_now += len(rows)

        trs = []
        for i, r in enumerate(rows):
            alt_bg = "#f8fafc" if i % 2 == 1 else "#ffffff"
            trs.append(f"""
              <tr>
                <td style="padding:10px 12px;border-top:1px solid #eef2f7;background:{alt_bg};font-weight:700;color:#0f172a;">
                  {r["name"]}
                </td>
                <td style="padding:10px 12px;border-top:1px solid #eef2f7;background:{alt_bg};white-space:nowrap;font-weight:700;color:{textc};">
                  {r["shift"]}
                </td>
              </tr>
            """)

        dept_color = DEPT_EMAIL_COLORS.get(dept, "#1e40af")

        dept_blocks.append(f"""
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                 style="margin-top:16px;border:1px solid #e6e6e6;border-radius:16px;overflow:hidden;background:#ffffff;">
            <tr>
              <td style="height:6px;background:{dept_color};font-size:0;line-height:0;">&nbsp;</td>
            </tr>
            <tr>
              <td style="padding:12px 14px;border-bottom:1px solid #eef2f7;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                  <tr>
                    <td style="font-size:16px;font-weight:900;color:{dept_color};">
                      {dept}
                    </td>
                    <td align="right">
                      <span style="
                        display:inline-block;
                        padding:6px 12px;
                        border-radius:12px;
                        font-size:13px;
                        font-weight:900;
                        color:{dept_color};
                        background:{dept_color}22;
                        border:1px solid {dept_color}55;
                      ">
                        TOTAL {len(rows)}
                      </span>
                    </td>
                  </tr>
                </table>
              </td>
            </tr>
            <tr>
              <td>
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                  <tr style="background:#f6f7f9;">
                    <th align="left" style="padding:10px 14px;border-bottom:1px solid #eef2f7;color:#334155;font-size:12px;letter-spacing:.4px;text-transform:uppercase;">
                      Employee
                    </th>
                    <th align="left" style="padding:10px 14px;border-bottom:1px solid #eef2f7;color:#334155;font-size:12px;letter-spacing:.4px;text-transform:uppercase;">
                      Status
                    </th>
                  </tr>
                  {''.join(trs)}
                </table>
              </td>
            </tr>
          </table>
        """)

    dept_html = "\n".join(dept_blocks) if dept_blocks else f"""
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-top:14px;">
        <tr>
          <td style="padding:14px;border-radius:14px;border:1px dashed rgba(15,23,42,.18);background:#ffffff;">
            <div style="font-weight:900;color:#334155;">No staff for current shift.</div>
            <div style="margin-top:6px;color:#64748b;font-size:13px;">Open the website for full details.</div>
          </td>
        </tr>
      </table>
    """

    pages_base = (pages_base or "").rstrip("/")

    html = f"""<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#eef1f7;font-family:Segoe UI,Arial,Helvetica,sans-serif;color:#0f172a;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#eef1f7;">
      <tr>
        <td align="center" style="padding:16px 10px;">
          <table role="presentation" width="680" cellpadding="0" cellspacing="0" style="max-width:680px;width:100%;">
            <tr>
              <td style="border-radius:20px;overflow:hidden;box-shadow:0 8px 28px rgba(30,64,175,.18);">

                <div style="background:linear-gradient(135deg,#1e40af 0%,#1976d2 50%,#0ea5e9 100%);padding:22px 18px;color:#fff;text-align:center;">
                  <div style="font-size:22px;font-weight:900;letter-spacing:-.2px;">üìã Duty Roster</div>
                  <div style="margin-top:8px;display:inline-block;background:rgba(255,255,255,.18);padding:6px 16px;border-radius:30px;font-size:13px;font-weight:700;">
                    üìÖ {date_label}
                  </div>
                </div>

                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#ffffff;">
                  <tr>
                    <td style="padding:16px 16px 10px 16px;">

                      <div style="margin:0 auto 12px auto;display:inline-block;padding:10px 14px;border-radius:14px;background:{bg};border:1px solid {border};color:{textc};font-weight:900;">
                        Current shift: {active_group}
                      </div>

                      <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:6px;">
                        <tr>
                          <td style="width:50%;padding-right:6px;">
                            <div style="border:1px solid rgba(15,23,42,.10);border-radius:14px;padding:10px 12px;text-align:center;background:#fff;">
                              <div style="font-size:22px;font-weight:900;color:#1e40af;">{total_now}</div>
                              <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.5px;text-transform:uppercase;">Now</div>
                            </div>
                          </td>
                          <td style="width:50%;padding-left:6px;">
                            <div style="border:1px solid rgba(15,23,42,.10);border-radius:14px;padding:10px 12px;text-align:center;background:#fff;">
                              <div style="font-size:22px;font-weight:900;color:#059669;">{depts_now}</div>
                              <div style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:.5px;text-transform:uppercase;">Departments</div>
                            </div>
                          </td>
                        </tr>
                      </table>

                      {dept_html}

                      <div style="text-align:center;margin-top:16px;">
                        <a href="{pages_base}/now/" style="display:inline-block;padding:12px 18px;border-radius:16px;background:linear-gradient(135deg,#1e40af,#1976d2);color:#fff;text-decoration:none;font-weight:900;">
                          Open Now Page
                        </a>
                        <span style="display:inline-block;width:10px;"></span>
                        <a href="{pages_base}/" style="display:inline-block;padding:12px 18px;border-radius:16px;background:#0ea5e9;color:#fff;text-decoration:none;font-weight:900;">
                          Open Full Page
                        </a>
                      </div>

                      <div style="margin-top:14px;text-align:center;color:#94a3b8;font-size:12px;line-height:1.9;">
                        Sent at <strong style="color:#64748b;">{sent_time}</strong> ¬∑ GitHub Actions
                      </div>

                    </td>
                  </tr>
                </table>

              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""





# =========================
# Email
# =========================


# =========================
# Helpers
# =========================
def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00A0", " ")).strip()

def to_western_digits(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    arabic = {'Ÿ†':'0','Ÿ°':'1','Ÿ¢':'2','Ÿ£':'3','Ÿ§':'4','Ÿ•':'5','Ÿ¶':'6','Ÿß':'7','Ÿ®':'8','Ÿ©':'9'}
    farsi  = {'€∞':'0','€±':'1','€≤':'2','€≥':'3','€¥':'4','€µ':'5','€∂':'6','€∑':'7','€∏':'8','€π':'9'}
    mp = {**arabic, **farsi}
    return "".join(mp.get(ch, ch) for ch in s)

def norm(s) -> str:
    return clean(to_western_digits(s))

def looks_like_time(s: str) -> bool:
    up = norm(s).upper()
    return bool(
        re.match(r"^\d{3,4}\s*H?\s*-\s*\d{3,4}\s*H?$", up)
        or re.match(r"^\d{3,4}\s*H$", up)
        or re.match(r"^\d{3,4}$", up)
    )

def looks_like_employee_name(s: str) -> bool:
    v = norm(s)
    if not v:
        return False
    up = v.upper()
    if looks_like_time(up):
        return False
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", up):
        return False
    # ŸÇŸàŸä: ÿßÿ≥ŸÖ - ÿ±ŸÇŸÖ
    if re.search(r"-\s*\d{3,}", v) and re.search(r"[A-Za-z\u0600-\u06FF]", v):
        return True
    # ÿ®ÿØŸäŸÑ: ŸÉŸÑŸÖÿ™ŸäŸÜ ÿ£Ÿà ÿ£ŸÉÿ´ÿ±
    parts = [p for p in v.split(" ") if p]
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", v) and len(parts) >= 2)

def looks_like_shift_code(s: str) -> bool:
    v = norm(s).upper()
    if not v:
        return False
    if looks_like_time(v):
        return False
    if v in ["OFF", "O", "LV", "TR", "ST", "SL", "AL", "STM", "STN"]:
        return True
    if re.match(r"^(MN|AN|NN|NT|ME|AE|NE)\d{1,2}", v):
        return True
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", v):
        return True
    return False

def map_shift(code: str):
    c0 = norm(code)
    c = c0.upper()
    if not c or c == "0":
        return ("-", "ÿ£ÿÆÿ±Ÿâ")

    if c == "AL" or "ANNUAL LEAVE" in c:
        return ("üèñÔ∏è Leave", "ÿ•ÿ¨ÿßÿ≤ÿßÿ™")
    if c == "SL" or "SICK LEAVE" in c:
        return ("ü§í Sick Leave", "ÿ•ÿ¨ÿßÿ≤ÿßÿ™")
    if c == "LV":
        return ("üèñÔ∏è Leave", "ÿ•ÿ¨ÿßÿ≤ÿßÿ™")
    if c in ["TR"] or "TRAINING" in c:
        return ("üìö Training", "ÿ™ÿØÿ±Ÿäÿ®")
    if c in ["ST", "STM", "STN"] or "STANDBY" in c:
        return ("üßç Standby", "ŸÖŸÜÿßŸàÿ®ÿßÿ™")
    if c in ["OFF", "O"] or re.search(r"(REST|OFF\s*DAY|REST\/OFF)", c):
        return ("üõå Off Day", "ÿ±ÿßÿ≠ÿ©")

    if c in SHIFT_MAP:
        return SHIFT_MAP[c]

    return (c0, "ÿ£ÿÆÿ±Ÿâ")

def current_shift_key(now: datetime) -> str:
    # 21:00‚Äì04:59 Night, 14:00‚Äì20:59 Afternoon, else Morning
    t = now.hour * 60 + now.minute
    if t >= 21 * 60 or t < 5 * 60:
        return "ŸÑŸäŸÑ"
    if t >= 14 * 60:
        return "ÿ∏Ÿáÿ±"
    return "ÿµÿ®ÿßÿ≠"


def roster_effective_datetime(now: datetime) -> datetime:
    """Night shift after midnight should still use yesterday's roster date (until 06:00 Muscat)."""
    active = current_shift_key(now)
    if active == "ŸÑŸäŸÑ" and now.hour < 6:
        return now - timedelta(days=1)
    return now

def download_excel(url: str) -> bytes:
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.content

def infer_pages_base_url():
    return "https://khalidsaif912.github.io/roster-site"


# =========================
# Detect rows/cols (Days row + Date numbers row)
# =========================
def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    count = 0
    for d in DAYS:
        if any(d in x for x in ups):
            count += 1
    return count

def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False

def find_days_and_dates_rows(ws, scan_rows: int = 80):
    """
    Ÿäÿ®ÿ≠ÿ´ ÿπŸÜ ÿµŸÅ ŸÅŸäŸá SUN..SAT ÿ®ŸÉÿ´ÿ±ÿ© ÿ´ŸÖ ÿµŸÅ ÿ™ÿ≠ÿ™Ÿá ŸÅŸäŸá ÿ£ÿ±ŸÇÿßŸÖ 1..31
    """
    max_r = min(ws.max_row, scan_rows)
    days_row = None

    for r in range(1, max_r + 1):
        vals = _row_values(ws, r)
        if _count_day_tokens(vals) >= 3:
            days_row = r
            break

    if not days_row:
        return None, None

    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        vals = _row_values(ws, r)
        nums = sum(1 for v in vals if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break

    return days_row, date_row

def find_day_col(ws, days_row: int, date_row: int, today_dow: int, today_day: int):
    """
    Ÿäÿ´ÿ®ÿ™ ÿßŸÑÿπŸÖŸàÿØ ÿßŸÑÿµÿ≠Ÿäÿ≠ ÿ®ÿßÿ≥ÿ™ÿÆÿØÿßŸÖ ÿßŸÑŸäŸàŸÖ + ÿ±ŸÇŸÖ ÿßŸÑÿ™ÿßÿ±ŸäÿÆ
    """
    if not days_row or not date_row:
        return None

    day_key = DAYS[today_dow]
    # Prefer (day + date) match
    for c in range(1, ws.max_column + 1):
        top = norm(ws.cell(row=days_row, column=c).value).upper()
        bot = norm(ws.cell(row=date_row, column=c).value)
        if day_key in top and _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    # Fallback: date-only
    for c in range(1, ws.max_column + 1):
        bot = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    return None

def find_employee_col(ws, start_row: int, max_scan_rows: int = 200):
    scores = {}
    r_end = min(ws.max_row, start_row + max_scan_rows)
    for r in range(start_row, r_end + 1):
        for c in range(1, ws.max_column + 1):
            if looks_like_employee_name(ws.cell(row=r, column=c).value):
                scores[c] = scores.get(c, 0) + 1
    if not scores:
        return None
    return max(scores.items(), key=lambda kv: kv[1])[0]


# =========================
# EXACT DESIGN (as you provided)
# =========================
CSS = r"""
    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê RESET ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    body {
      margin:0; padding:0;
      background:#eef1f7;
      font-family:'Segoe UI', system-ui, -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif;
      color:#0f172a;
      -webkit-font-smoothing:antialiased;
    }
    * { box-sizing:border-box; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê WRAP ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .wrap { max-width:680px; margin:0 auto; padding:16px 14px 28px; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê HEADER ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .header {
      background:linear-gradient(135deg, #1e40af 0%, #1976d2 50%, #0ea5e9 100%);
      color:#fff;
      padding:26px 18px 24px;
      border-radius:20px;
      text-align:center;
      box-shadow:0 8px 28px rgba(30,64,175,.25);
      position:relative;
      overflow:hidden;
    }
    .header::before {
      content:''; position:absolute;
      top:-30px; right:-40px;
      width:140px; height:140px;
      border-radius:50%;
      background:rgba(255,255,255,.08);
    }
    .header::after {
      content:''; position:absolute;
      bottom:-50px; left:-30px;
      width:160px; height:160px;
      border-radius:50%;
      background:rgba(255,255,255,.06);
    }
    .header h1 { margin:0; font-size:24px; font-weight:800; position:relative; z-index:1; letter-spacing:-.3px; }
    .header .dateTag {
      display:inline-block; margin-top:10px;
      background:rgba(255,255,255,.18);
      padding:5px 18px; border-radius:30px;
      font-size:13px; font-weight:600; letter-spacing:.3px;
      position:relative; z-index:1;
    }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê SUMMARY BAR ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .summaryBar { display:flex; justify-content:center; gap:12px; margin-top:14px; }
    .summaryChip {
      background:#fff;
      border:1px solid rgba(15,23,42,.1);
      border-radius:14px;
      padding:10px 20px;
      text-align:center;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
    }
    .summaryChip .chipVal { font-size:22px; font-weight:900; color:#1e40af; }
    .summaryChip .chipLabel { font-size:11px; font-weight:600; color:#64748b; text-transform:uppercase; letter-spacing:.6px; margin-top:2px; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê DEPARTMENT CARD ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .deptCard {
      margin-top:18px;
      background:#fff;
      border-radius:18px;
      overflow:hidden;
      border:1px solid rgba(15,23,42,.07);
      box-shadow:0 4px 18px rgba(15,23,42,.08);
    }
    .deptHead {
      display:flex;
      align-items:center;
      gap:12px;
      padding:14px 16px;
      background:#fff;
    }
    .deptIcon {
      width:40px; height:40px;
      border-radius:12px;
      display:flex; align-items:center; justify-content:center;
      flex-shrink:0;
    }
    .deptTitle { font-size:18px; font-weight:800; color:#1e293b; flex:1; letter-spacing:-.2px; }
    .deptBadge { min-width:48px; padding:6px 10px; border-radius:12px; text-align:center; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê SHIFT STACK ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .shiftStack { padding:10px; display:flex; flex-direction:column; gap:8px; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê SHIFT CARD ‚Äî <details> ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .shiftCard {
      border-radius:14px;
      overflow:hidden;
    }

    .shiftSummary {
      display:flex;
      align-items:center;
      gap:10px;
      padding:11px 14px;
      cursor:pointer;
      list-style:none;
      -webkit-appearance:none;
      appearance:none;
      user-select:none;
    }
    .shiftSummary::-webkit-details-marker { display:none; }
    .shiftSummary::marker              { display:none; }

    .shiftIcon  { font-size:20px; line-height:1; flex-shrink:0; }
    .shiftLabel { font-size:15px; font-weight:800; flex:1; letter-spacing:-.1px; }
    .shiftCount {
      font-size:13px; font-weight:800;
      padding:3px 10px; border-radius:20px;
      flex-shrink:0;
    }

    /* chevron ŸäÿØŸàÿ± ŸÑŸÖÿß ŸäŸÅÿ™ÿ≠ */
    .shiftSummary::after {
      content:'‚ñæ';
      font-size:14px;
      color:#94a3b8;
      transition:transform .2s;
      flex-shrink:0;
    }
    .shiftCard[open] .shiftSummary::after {
      transform:rotate(180deg);
    }

    .shiftBody { background:rgba(255,255,255,.7); }

    /* ‚îÄ‚îÄ employee row ‚îÄ‚îÄ */
    .empRow {
      display:flex;
      align-items:center;
      justify-content:space-between;
      padding:9px 16px;
      border-top:1px solid rgba(15,23,42,.06);
    }
    .empRowAlt { background:rgba(15,23,42,.02); }
    .empName  { font-size:15px; font-weight:700; color:#1e293b; }
    .empStatus { font-size:13px; font-weight:600; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê CTA ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .btnWrap { margin-top:20px; text-align:center; }
    .btn {
      display:inline-block;
      padding:14px 38px;
      border-radius:16px;
      background:linear-gradient(135deg, #1e40af, #1976d2);
      color:#fff !important;
      text-decoration:none;
      font-weight:800;
      font-size:15px;
      box-shadow:0 6px 20px rgba(30,64,175,.3);
    }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê FOOTER ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    .footer { margin-top:18px; text-align:center; font-size:12px; color:#94a3b8; padding:12px 0; line-height:1.9; }
    .footer strong { color:#64748b; }

    /* ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê MOBILE ‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê */
    @media (max-width:480px){
      .wrap            { padding:12px 10px 22px; }
      .header h1       { font-size:21px; }
      .deptTitle       { font-size:16px; }
      .empName         { font-size:14px; }
      .empStatus       { font-size:12px; }
      .shiftLabel      { font-size:14px; }
      .summaryBar      { gap:8px; }
      .summaryChip     { padding:8px 14px; }
      .summaryChip .chipVal { font-size:19px; }
    }
"""

DEPT_COLORS = ["#2563eb", "#7c3aed", "#0891b2", "#059669", "#dc2626", "#ea580c"]

SVG_ICON = """
<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>
"""

def shift_style(grp: str, label_text: str):
    """
    Returns: (shift_title, icon, border_color, bg_color, text_color, count_bg)
    """
    if grp == "ÿµÿ®ÿßÿ≠":
        return ("Morning", "‚òÄÔ∏è", "#f59e0b44", "#fef3c7", "#92400e", "#f59e0b22")
    if grp == "ÿ∏Ÿáÿ±":
        return ("Afternoon", "üå§Ô∏è", "#f9731644", "#ffedd5", "#9a3412", "#f9731622")
    if grp == "ŸÑŸäŸÑ":
        return ("Night", "üåô", "#8b5cf644", "#ede9fe", "#5b21b6", "#8b5cf622")
    if grp == "ÿ±ÿßÿ≠ÿ©":
        return ("Off Day", "üõãÔ∏è", "#6366f144", "#e0e7ff", "#3730a3", "#6366f122")
    if grp == "ÿ•ÿ¨ÿßÿ≤ÿßÿ™":
        # differentiate sick via label
        if "SICK" in label_text.upper() or "ü§í" in label_text:
            return ("Sick Leave", "üè•", "#ef444444", "#fee2e2", "#991b1b", "#ef444422")
        return ("Annual Leave", "‚úàÔ∏è", "#10b98144", "#d1fae5", "#065f46", "#10b98122")
    if grp == "ÿ™ÿØÿ±Ÿäÿ®":
        return ("Training", "üìö", "#0ea5e944", "#e0f2fe", "#075985", "#0ea5e922")
    if grp == "ŸÖŸÜÿßŸàÿ®ÿßÿ™":
        return ("Standby", "üßç", "#94a3b844", "#f1f5f9", "#334155", "#94a3b822")
    return ("Other", "üìå", "#64748b44", "#f8fafc", "#334155", "#64748b22")

def dept_card_html(dept_name: str, dept_color: str, buckets: dict, open_group: str | None = None):
    total = sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
    shift_blocks = []

    for g in GROUP_ORDER:
        rows = buckets.get(g, [])
        if not rows:
            continue

        # use first row label for style decision (sick vs annual)
        first_label = rows[0]["shift"] if rows else ""
        title, icon, border, bg, text_color, count_bg = shift_style(g, first_label)

        # open only one group if requested
        open_attr = " open" if (open_group and g == open_group) else ""

        emp_rows_html = []
        for i, x in enumerate(rows):
            alt = " empRowAlt" if i % 2 == 1 else ""
            emp_rows_html.append(
                f"""<div class="empRow{alt}">
       <span class="empName">{x["name"]}</span>
       <span class="empStatus" style="color:{text_color};">{x["shift"]}</span>
     </div>"""
            )

        shift_blocks.append(
            f"""
    <details class="shiftCard" style="border:1px solid {border}; background:{bg};"{open_attr}>
      <summary class="shiftSummary" style="background:{bg}; border-bottom:1px solid {border.replace('44','33')};">
        <span class="shiftIcon">{icon}</span>
        <span class="shiftLabel" style="color:{text_color};">{title}</span>
        <span class="shiftCount" style="background:{count_bg}; color:{text_color};">{len(rows)}</span>
      </summary>
      <div class="shiftBody">
        {''.join(emp_rows_html)}
      </div>
    </details>
            """
        )

    if not shift_blocks:
        shift_blocks_html = '<div class="shiftStack"><div class="footer" style="margin:0; padding:14px 0;">No data for today</div></div>'
    else:
        shift_blocks_html = f'<div class="shiftStack">{"".join(shift_blocks)}</div>'

    return f"""
    <div class="deptCard">
      <div style="height:5px; background:linear-gradient(to right, {dept_color}, {dept_color}cc);"></div>

      <div class="deptHead" style="border-bottom:2px solid {dept_color}18;">
        <div class="deptIcon" style="background:{dept_color}15; color:{dept_color};">
          {SVG_ICON}
        </div>
        <div class="deptTitle">{dept_name}</div>
        <div class="deptBadge" style="background:{dept_color}12; color:{dept_color}; border:1px solid {dept_color}28;">
          <span style="font-size:10px;opacity:.7;display:block;margin-bottom:1px;text-transform:uppercase;letter-spacing:.5px;">Total</span>
          <span style="font-size:17px;font-weight:900;">{total}</span>
        </div>
      </div>

      {shift_blocks_html}
    </div>
    """

def page_shell_html(date_label: str, employees_total: int, departments_total: int, dept_cards_html: str, cta_url: str, sent_time: str):
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster</title>
  <style>
{CSS}
  </style>
</head>
<body>
<div class="wrap">

  <!-- ‚ïê‚ïê‚ïê‚ïê HEADER ‚ïê‚ïê‚ïê‚ïê -->
  <div class="header">
    <h1>üìã Duty Roster</h1>
    <div class="dateTag">üìÖ {date_label}</div>
  </div>

  <!-- ‚ïê‚ïê‚ïê‚ïê SUMMARY CHIPS ‚ïê‚ïê‚ïê‚ïê -->
  <div class="summaryBar">
    <div class="summaryChip">
      <div class="chipVal">{employees_total}</div>
      <div class="chipLabel">Employees</div>
    </div>
    <div class="summaryChip">
      <div class="chipVal" style="color:#059669;">{departments_total}</div>
      <div class="chipLabel">Departments</div>
    </div>
  </div>

  <!-- ‚ïê‚ïê‚ïê‚ïê DEPARTMENT CARDS ‚ïê‚ïê‚ïê‚ïê -->
  {dept_cards_html}

  <!-- ‚ïê‚ïê‚ïê‚ïê CTA ‚ïê‚ïê‚ïê‚ïê -->
  <div class="btnWrap">
    <a class="btn" href="{cta_url}">üìã View Full Duty Roster</a>
  </div>

  <!-- ‚ïê‚ïê‚ïê‚ïê FOOTER ‚ïê‚ïê‚ïê‚ïê -->

  <div style="margin-top:18px;background:#fff;border-radius:18px;border:1px solid rgba(15,23,42,.07);box-shadow:0 4px 18px rgba(15,23,42,.08);padding:14px;text-align:center;">
    <div style="font-weight:900;font-size:16px;color:#0f172a;margin-bottom:10px;">üì© Subscribe</div>
    <div style="color:#64748b;font-weight:700;font-size:13px;margin-bottom:10px;">Enter your email to receive roster updates automatically</div>
    <div style="display:flex;gap:8px;justify-content:center;flex-wrap:wrap;">
      <input id="subEmail" type="email" placeholder="name@example.com"
        style="padding:12px 14px;border-radius:14px;border:1px solid rgba(15,23,42,.12);min-width:240px;font-weight:800;outline:none;">
      <button id="subBtn"
        style="padding:12px 16px;border-radius:14px;border:none;background:linear-gradient(135deg,#1e40af,#1976d2);color:#fff;font-weight:900;cursor:pointer;">
        Subscribe
      </button>
    </div>
    <div id="subMsg" style="margin-top:10px;font-weight:900;"></div>
    <div style="margin-top:8px;font-size:12px;color:#94a3b8;">We won‚Äôt share your email.</div>
  </div>


  <div class="footer">
    Sent at <strong>{sent_time}</strong>
     &nbsp;¬∑&nbsp; Total: <strong>{employees_total} employees</strong>
  </div>

</div>
</body>
</html>"""


# =========================
# Email
# =========================

def send_email(subject: str, html: str, mail_to: str):
    """Send HTML email to one or more recipients (comma-separated)."""
    mail_to = (mail_to or "").strip()
    recipients = [x.strip() for x in mail_to.split(",") if x.strip()]
    if not recipients:
        raise RuntimeError("MAIL_TO is empty (no recipients). Add MAIL_TO secret and/or subscribers.")

    msg = MIMEText(html, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = MAIL_FROM
    msg["To"] = ", ".join(recipients)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.sendmail(MAIL_FROM, recipients, msg.as_string())

def build_cards_for_date(wb, target_date: datetime, active_group: str):
    """Build department cards HTML + totals for a given date."""
    today_dow = (target_date.weekday() + 1) % 7  # Sun=0..Sat=6
    today_day = target_date.day

    dept_cards = []
    employees_total = 0
    depts_count = 0

    for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        day_col = find_day_col(ws, days_row, date_row, today_dow, today_day)

        if not (days_row and date_row and day_col):
            continue

        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        if not emp_col:
            continue

        buckets = {k: [] for k in GROUP_ORDER}

        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue

            raw = norm(ws.cell(row=r, column=day_col).value)
            if not looks_like_shift_code(raw):
                continue

            label, grp = map_shift(raw)
            buckets.setdefault(grp, []).append({"name": name, "shift": label})

        dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]
        open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None
        dept_cards.append(dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full))

        employees_total += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
        depts_count += 1

    return "\n".join(dept_cards), employees_total, depts_count




def build_rows_for_email(wb, target_date: datetime, active_group: str):
    """Return rows grouped by department for EMAIL (active shift only)."""
    today_dow = (target_date.weekday() + 1) % 7  # Sun=0..Sat=6
    today_day = target_date.day

    rows_by_dept = []
    for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        day_col = find_day_col(ws, days_row, date_row, today_dow, today_day)
        if not (days_row and date_row and day_col):
            continue

        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        if not emp_col:
            continue

        rows = []
        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue

            raw = norm(ws.cell(row=r, column=day_col).value)
            if not looks_like_shift_code(raw):
                continue

            label, grp = map_shift(raw)
            if grp == active_group:
                rows.append({"name": name, "shift": label})

        if not rows:
            continue

        dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]
        base_color = dept_color.get("base") if isinstance(dept_color, dict) else str(dept_color)
        rows_by_dept.append({"dept": dept_name, "color": base_color or "#2563eb", "rows": rows})

    return rows_by_dept


def page_shell_html_full_with_picker(month_iso: str, employees_total: int, departments_total: int, dept_cards_html: str, cta_url: str, sent_time: str):
    """Full page with a date picker. It loads per-day HTML from ./data/roster.json."""
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster</title>
  <style>
{CSS}
    /* Date picker inside header */
    .datePicker {{
      margin-top:10px;
      background:rgba(255,255,255,.18);
      padding:6px 14px;
      border-radius:30px;
      border:1px solid rgba(255,255,255,.30);
      color:#fff;
      font-size:13px;
      font-weight:700;
      outline:none;
    }}
    .datePicker::-webkit-calendar-picker-indicator {{
      filter: invert(1);
      opacity: .9;
      cursor: pointer;
    }}
  </style>
</head>
<body>
<div class="wrap">

  <div class="header">
    <h1>üìã Duty Roster</h1>
    <input id="datePicker" class="datePicker" type="date" aria-label="Choose date">
  </div>

  <div class="summaryBar">
    <div class="summaryChip">
      <div id="empCount" class="chipVal">{employees_total}</div>
      <div class="chipLabel">Employees</div>
    </div>
    <div class="summaryChip">
      <div id="deptCount" class="chipVal" style="color:#059669;">{departments_total}</div>
      <div class="chipLabel">Departments</div>
    </div>
  </div>

  <div id="content">
    {dept_cards_html}
  </div>

  <div class="btnWrap">
    <a class="btn" href="{cta_url}">üåô View NOW</a>
  </div>

  <div class="footer">
    Sent at <strong>{sent_time}</strong>
     &nbsp;¬∑&nbsp; Month: <strong>{month_iso}</strong>
  </div>

</div>

<script>
(async function() {{
  const picker = document.getElementById('datePicker');
  const content = document.getElementById('content');
  const empCount = document.getElementById('empCount');

  const res = await fetch('./data/roster.json', {{ cache: 'no-store' }});
  const data = await res.json();

  const days = Object.keys(data.days || {{}}).sort();
  if (!days.length) return;

  const urlDate = new URLSearchParams(location.search).get('date');
  const defaultDay = (urlDate && data.days[urlDate]) ? urlDate : (data.default_day || days[0]);

  picker.min = days[0];
  picker.max = days[days.length - 1];
  picker.value = defaultDay;

  function render(dayISO) {{
    const d = data.days[dayISO];
    if (!d) {{
      content.innerHTML = '<div class="deptCard" style="padding:16px;text-align:center;">No data for this date.</div>';
      empCount.textContent = '0';
      return;
    }}
    content.innerHTML = d.cards_html || '';
    empCount.textContent = String(d.employees_total || 0);
    history.replaceState({{}}, '', '?date=' + dayISO);
  }}

  render(defaultDay);
  picker.addEventListener('change', (e) => render(e.target.value));
}})();
</script>

</body>
</html>"""

def load_subscribers() -> list[str]:
    """Load subscriber emails from Google Apps Script (Google Sheet) via GET ?token=...

    Expected JSON: {"ok": true, "emails": ["a@b.com", ...]}
    """
    if not SUBSCRIBE_URL or not SUBSCRIBE_TOKEN:
        return []
    try:
        r = requests.get(SUBSCRIBE_URL, params={"token": SUBSCRIBE_TOKEN}, timeout=30)
        r.raise_for_status()
        j = r.json()
        if j.get("ok"):
            emails = []
            for e in (j.get("emails") or []):
                e = str(e).strip().lower()
                if e and "@" in e:
                    emails.append(e)
            # unique preserve order
            seen=set()
            out=[]
            for e in emails:
                if e not in seen:
                    seen.add(e)
                    out.append(e)
            return out
    except Exception:
        return []
    return []



def main():
    if not EXCEL_URL:
        raise RuntimeError("EXCEL_URL missing")

    now = datetime.now(TZ)
    effective = roster_effective_datetime(now)
    # Sun=0..Sat=6 (based on roster effective date)
    today_dow = (effective.weekday() + 1) % 7
    today_day = effective.day

    active_group = current_shift_key(now)  # "ÿµÿ®ÿßÿ≠" / "ÿ∏Ÿáÿ±" / "ŸÑŸäŸÑ"
    pages_base = (PAGES_BASE_URL or infer_pages_base_url()).rstrip("/")

    data = download_excel(EXCEL_URL)
    wb = load_workbook(BytesIO(data), data_only=True)

    dept_cards_all = []
    dept_cards_now = []
    employees_total_all = 0
    employees_total_now = 0
    depts_count = 0

    for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        day_col = find_day_col(ws, days_row, date_row, today_dow, today_day)

        if not (days_row and date_row and day_col):
            # skip if sheet layout unexpected
            continue

        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        if not emp_col:
            continue

        buckets = {k: [] for k in GROUP_ORDER}
        buckets_now = {k: [] for k in GROUP_ORDER}

        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue

            raw = norm(ws.cell(row=r, column=day_col).value)
            if not looks_like_shift_code(raw):
                continue

            label, grp = map_shift(raw)
            buckets.setdefault(grp, []).append({"name": name, "shift": label})

            if grp == active_group:
                buckets_now.setdefault(grp, []).append({"name": name, "shift": label})

        dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]
        card_all = dept_card_html(dept_name, dept_color, buckets, open_group=None)
        dept_cards_all.append(card_all)

        # For NOW page: open the active shift group by default
        card_now = dept_card_html(dept_name, dept_color, buckets_now, open_group=active_group)
        dept_cards_now.append(card_now)

        employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
        employees_total_now += sum(len(buckets_now.get(g, [])) for g in GROUP_ORDER)

        depts_count += 1

    
# pages
    os.makedirs("docs", exist_ok=True)
    os.makedirs("docs/now", exist_ok=True)
    os.makedirs("docs/data", exist_ok=True)

    # Display date based on roster effective date
    try:
        date_label = effective.strftime("%-d %B %Y")
    except Exception:
        date_label = effective.strftime("%d %B %Y")

    sent_time = now.strftime("%H:%M")

    full_url = f"{pages_base}/"
    now_url = f"{pages_base}/now/"

    # Build per-day HTML for the whole month (static JSON for the date picker)
    month_iso = effective.strftime("%Y-%m")
    year = effective.year
    month = effective.month
    last_day = calendar.monthrange(year, month)[1]

    roster_days = {}
    for d in range(1, last_day + 1):
        dt = effective.replace(day=d)
        cards_html, emp_total, dept_total = build_cards_for_date(wb, dt, active_group)
        day_iso = dt.strftime("%Y-%m-%d")
        roster_days[day_iso] = {
            "cards_html": cards_html,
            "employees_total": emp_total,
            "departments_total": dept_total,
        }

    default_day = effective.strftime("%Y-%m-%d")
    with open("docs/data/roster.json", "w", encoding="utf-8") as f:
        json.dump({"month": month_iso, "default_day": default_day, "days": roster_days}, f, ensure_ascii=False)

    # Full page with picker (initially renders default day)
    cards_today, emp_today, dept_today = build_cards_for_date(wb, effective, active_group)
    html_full = page_shell_html_full_with_picker(
        month_iso=month_iso,
        employees_total=emp_today,
        departments_total=dept_today,
        dept_cards_html=cards_today,
        cta_url=now_url,
        sent_time=sent_time,
    )

    html_now = page_shell_html(
        date_label=date_label,
        employees_total=employees_total_now,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_now),
        cta_url=full_url,  # button on now page goes to FULL page
        sent_time=sent_time,
    )

    with open("docs/index.html", "w", encoding="utf-8") as f:
        f.write(html_full)

    with open("docs/now/index.html", "w", encoding="utf-8") as f:
        f.write(html_now)

    # Email: send NOW page design (same exact template)
    subject = f"Duty Roster ‚Äî {active_group} ‚Äî {effective.strftime('%Y-%m-%d')}"
    rows_by_dept = build_rows_for_email(wb, effective, active_group)
    email_html = build_pretty_email_html(active_group, now, rows_by_dept, pages_base)

    # ‚úÖ Merge manual recipients (MAIL_TO) + subscribers from Google Sheet
    manual = [x.strip().lower() for x in MAIL_TO.split(",") if x.strip()]
    subs = load_subscribers()
    all_recipients = sorted(set(manual + subs))
    mail_to_final = ",".join(all_recipients)

    send_email(subject, email_html, mail_to_final)

if __name__ == "__main__":
    main()