#!/usr/bin/env python3
"""
generate_employee_schedules.py

Ø³ÙƒØ±ÙŠØ¨Øª Ù…Ø³ØªÙ‚Ù„ Ù„ØªÙˆÙ„ÙŠØ¯ Ù…Ù„ÙØ§Øª JSON Ù„Ø¬Ø¯Ø§ÙˆÙ„ Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†
ÙŠØ¹Ù…Ù„ Ø¨Ø´ÙƒÙ„ Ù…Ù†ÙØµÙ„ Ø¹Ù† generate_and_send.py

Ø§Ù„Ø§Ø³ØªØ®Ø¯Ø§Ù…:
    python generate_employee_schedules.py
    python generate_employee_schedules.py --month 2026-03
"""

import os
import re
import json
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO
from collections import defaultdict

import requests
from openpyxl import load_workbook


# =========================
# Settings
# =========================
EXCEL_URL = os.environ.get("EXCEL_URL", "").strip()
TZ = ZoneInfo("Asia/Muscat")

DEPARTMENTS = [
    ("Officers", "Officers"),
    ("Supervisors", "Supervisors"),
    ("Load Control", "Load Control"),
    ("Export Checker", "Export Checker"),
    ("Export Operators", "Export Operators"),
    ("Unassigned", "Unassigned"),
]

DAYS = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

SHIFT_MAP = {
    "MN06": ("ðŸŒ… Morning (MN06)", "Morning"),
    "ME06": ("ðŸŒ… Morning (ME06)", "Morning"),
    "ME07": ("ðŸŒ… Morning (ME07)", "Morning"),
    "MN12": ("ðŸŒ† Afternoon (MN12)", "Afternoon"),
    "AN13": ("ðŸŒ† Afternoon (AN13)", "Afternoon"),
    "AE14": ("ðŸŒ† Afternoon (AE14)", "Afternoon"),
    "NN21": ("ðŸŒ™ Night (NN21)", "Night"),
    "NE22": ("ðŸŒ™ Night (NE22)", "Night"),
}


# =========================
# Helper Functions
# =========================
def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00A0", " ")).strip()


def to_western_digits(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    arabic = {'Ù ':'0','Ù¡':'1','Ù¢':'2','Ù£':'3','Ù¤':'4','Ù¥':'5','Ù¦':'6','Ù§':'7','Ù¨':'8','Ù©':'9'}
    farsi = {'Û°':'0','Û±':'1','Û²':'2','Û³':'3','Û´':'4','Ûµ':'5','Û¶':'6','Û·':'7','Û¸':'8','Û¹':'9'}
    mp = {**arabic, **farsi}
    return "".join(mp.get(ch, ch) for ch in s)


def norm(s) -> str:
    return clean(to_western_digits(s))


def looks_like_time(s: str) -> bool:
    up = norm(s).upper()
    return bool(
        re.match(r"^\d{3,4}\s*H?\s*-\s*\d{3,4}\s*H?$", up)
        or re.match(r"^\d{3,4}\s*H$", up)
        or re.match(r"^\d{3,4}$", up)
    )


def looks_like_employee_name(s: str) -> bool:
    v = norm(s)
    if not v:
        return False
    up = v.upper()
    if looks_like_time(up):
        return False
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", up):
        return False
    if re.search(r"-\s*\d{3,}", v) and re.search(r"[A-Za-z\u0600-\u06FF]", v):
        return True
    parts = [p for p in v.split(" ") if p]
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", v) and len(parts) >= 2)


def looks_like_shift_code(s: str) -> bool:
    v = norm(s).upper()
    if not v:
        return False
    if looks_like_time(v):
        return False
    if v in ["OFF", "O", "LV", "TR", "ST", "SL", "AL", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14", "OT"]:
        return True
    if re.match(r"^(MN|AN|NN|NT|ME|AE|NE)\d{1,2}", v):
        return True
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", v):
        return True
    if len(v) >= 3 and re.search(r"[A-Z]", v):
        return True
    return False


def map_shift(code: str):
    c0 = norm(code)
    c = c0.upper()
    if not c or c == "0":
        return ("-", "Other")
    
    if c == "AL" or c == "LV" or "ANNUAL LEAVE" in c:
        return ("âœˆï¸ Annual Leave", "Annual Leave")
    
    if c == "SL" or "SICK LEAVE" in c:
        return ("ðŸ¤’ Sick Leave", "Sick Leave")
    
    if c in ["TR"] or "TRAINING" in c:
        return ("ðŸ“š Training", "Training")
    
    if c in ["ST", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14"] or "STANDBY" in c:
        return (f"ðŸ§ {c0}", "Standby")
    
    if c == "OT" or c.startswith("OT"):
        return (f"â±ï¸ {c0}", "Standby")
    
    if c in ["OFF", "O"] or re.search(r"(REST|OFF\s*DAY|REST\/OFF)", c):
        return ("ðŸ›Œ Off Day", "Off Day")
    
    if c in SHIFT_MAP:
        return SHIFT_MAP[c]
    
    return (f"â“ {c0}", "Other")


def download_excel(url: str) -> bytes:
    print(f"ðŸ“¥ Downloading Excel from: {url[:50]}...")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    print("âœ… Excel downloaded successfully")
    return r.content


def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]


def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    count = 0
    for d in DAYS:
        if any(d in x for x in ups):
            count += 1
    return count


def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False


def find_days_and_dates_rows(ws, scan_rows: int = 80):
    max_r = min(ws.max_row, scan_rows)
    days_row = None
    
    for r in range(1, max_r + 1):
        vals = _row_values(ws, r)
        if _count_day_tokens(vals) >= 3:
            days_row = r
            break
    
    if not days_row:
        return None, None
    
    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        vals = _row_values(ws, r)
        nums = sum(1 for v in vals if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break
    
    return days_row, date_row


def find_employee_col(ws, start_row: int):
    for c in range(1, min(10, ws.max_column + 1)):
        val = norm(ws.cell(row=start_row, column=c).value)
        if looks_like_employee_name(val):
            return c
    return None


def get_daynum_to_col(ws, date_row: int):
    daynum_to_col = {}
    for c in range(1, ws.max_column + 1):
        val = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(val):
            daynum_to_col[int(float(val))] = c
    return daynum_to_col


def extract_employee_id(name_str):
    """ÙŠØ³ØªØ®Ø±Ø¬ Ø§Ù„Ø±Ù‚Ù… Ù…Ù† Ù†Øµ Ù…Ø«Ù„: Ahmed Ali - 12345"""
    match = re.search(r'-\s*(\d+)\s*$', name_str)
    if match:
        return match.group(1).strip()
    return None


# =========================
# Main Functions
# =========================
def generate_employee_schedules(wb, year: int, month: int):
    """
    ØªÙˆÙ„ÙŠØ¯ Ù…Ù„ÙØ§Øª JSON Ù„Ø¬Ø¯Ø§ÙˆÙ„ ÙƒÙ„ Ù…ÙˆØ¸Ù
    """
    print(f"\nðŸ“… Generating employee schedules for {year}-{month:02d}...")
    
    all_employees = defaultdict(lambda: {
        "name": "",
        "id": "",
        "department": "",
        "schedules": {}
    })
    
    # Ù…Ø¹Ø§Ù„Ø¬Ø© ÙƒÙ„ Ù‚Ø³Ù…
    for sheet_name, dept_name in DEPARTMENTS:
        if sheet_name not in wb.sheetnames:
            continue
        
        print(f"  ðŸ“‹ Processing {dept_name}...")
        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        
        if not (days_row and date_row):
            print(f"    âš ï¸  Could not find days/dates rows")
            continue
        
        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        daynum_to_col = get_daynum_to_col(ws, date_row)
        
        if not emp_col:
            print(f"    âš ï¸  Could not find employee column")
            continue
        
        emp_count = 0
        # Ù…Ø¹Ø§Ù„Ø¬Ø© ÙƒÙ„ Ù…ÙˆØ¸Ù
        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue
            
            emp_id = extract_employee_id(name)
            if not emp_id:
                continue
            
            emp_name = re.sub(r'\s*-\s*\d+\s*$', '', name).strip()
            
            # Ù‚Ø±Ø§Ø¡Ø© Ù…Ù†Ø§ÙˆØ¨Ø§Øª Ø§Ù„Ø´Ù‡Ø±
            month_schedule = []
            for day_num in sorted(daynum_to_col.keys()):
                col = daynum_to_col[day_num]
                raw = norm(ws.cell(row=r, column=col).value)
                
                if looks_like_shift_code(raw):
                    label, group = map_shift(raw)
                    
                    try:
                        date_obj = datetime(year, month, day_num, tzinfo=TZ)
                        day_name_ar = ["Ø§Ù„Ø£Ø­Ø¯", "Ø§Ù„Ø§Ø«Ù†ÙŠÙ†", "Ø§Ù„Ø«Ù„Ø§Ø«Ø§Ø¡", "Ø§Ù„Ø£Ø±Ø¨Ø¹Ø§Ø¡", "Ø§Ù„Ø®Ù…ÙŠØ³", "Ø§Ù„Ø¬Ù…Ø¹Ø©", "Ø§Ù„Ø³Ø¨Øª"]
                        day_name_en = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
                        dow = (date_obj.weekday() + 1) % 7
                        
                        month_schedule.append({
                            "date": date_obj.strftime("%Y-%m-%d"),
                            "day": day_num,
                            "day_name_ar": day_name_ar[dow],
                            "day_name_en": day_name_en[dow],
                            "shift_code": raw.upper(),
                            "shift_label": label,
                            "shift_group": group
                        })
                    except ValueError:
                        continue
            
            if month_schedule:
                month_key = f"{year}-{month:02d}"
                all_employees[emp_id]["name"] = emp_name
                all_employees[emp_id]["id"] = emp_id
                all_employees[emp_id]["department"] = dept_name
                all_employees[emp_id]["schedules"][month_key] = month_schedule
                emp_count += 1
        
        print(f"    âœ… Processed {emp_count} employees")
    
    # Ø­ÙØ¸ Ù…Ù„ÙØ§Øª JSON
    schedules_dir = "docs/schedules"
    os.makedirs(schedules_dir, exist_ok=True)
    
    saved_count = 0
    for emp_id, data in all_employees.items():
        filepath = f"{schedules_dir}/{emp_id}.json"
        
        # Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù‚Ø¯ÙŠÙ…Ø©
        existing_data = {"name": "", "id": emp_id, "department": "", "schedules": {}}
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except:
                pass
        
        # Ø¯Ù…Ø¬ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
        existing_data["name"] = data["name"]
        existing_data["department"] = data["department"]
        existing_data["schedules"].update(data["schedules"])
        
        # Ø­ÙØ¸
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        
        saved_count += 1
    
    print(f"\nâœ… Generated schedules for {saved_count} employees")
    return saved_count


def generate_schedule_index():
    """
    ÙŠÙ†Ø´Ø¦ Ù…Ù„Ù index Ù„Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…ÙˆØ¸ÙÙŠÙ†
    """
    print("\nðŸ“‘ Generating index...")
    schedules_dir = "docs/schedules"
    
    if not os.path.exists(schedules_dir):
        print("  âš ï¸  Schedules directory not found")
        return
    
    employees_list = []
    
    for filename in os.listdir(schedules_dir):
        if filename.endswith('.json') and filename != 'index.json':
            emp_id = filename.replace('.json', '')
            filepath = os.path.join(schedules_dir, filename)
            
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    employees_list.append({
                        "id": emp_id,
                        "name": data.get("name", ""),
                        "department": data.get("department", ""),
                        "months": sorted(list(data.get("schedules", {}).keys()))
                    })
            except:
                continue
    
    employees_list.sort(key=lambda x: (x["department"], x["name"]))
    
    index_file = os.path.join(schedules_dir, "index.json")
    with open(index_file, 'w', encoding='utf-8') as f:
        json.dump({
            "total": len(employees_list),
            "employees": employees_list,
            "last_updated": datetime.now(TZ).isoformat()
        }, f, ensure_ascii=False, indent=2)
    
    print(f"âœ… Generated index for {len(employees_list)} employees")


# =========================
# Main
# =========================
def add_months(year, month, delta):
    """Ø¥Ø¶Ø§ÙØ©/Ø·Ø±Ø­ Ø£Ø´Ù‡Ø± Ù…Ø¹ Ù…Ø¹Ø§Ù„Ø¬Ø© ØªØ¬Ø§ÙˆØ² Ø§Ù„Ø³Ù†Ø©"""
    month += delta
    while month > 12:
        month -= 12
        year += 1
    while month < 1:
        month += 12
        year -= 1
    return year, month


def detect_month_from_url(url: str):
    """Ø§Ø³ØªÙ†ØªØ§Ø¬ Ø§Ù„Ø´Ù‡Ø± Ù…Ù† Ø§Ø³Ù… Ù…Ù„Ù Excel ÙÙŠ Ø§Ù„Ø±Ø§Ø¨Ø·"""
    if not url:
        return None
    month_map = {
        "jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,
        "apr":4,"april":4,"may":5,"jun":6,"june":6,"jul":7,"july":7,
        "aug":8,"august":8,"sep":9,"sept":9,"september":9,
        "oct":10,"october":10,"nov":11,"november":11,"dec":12,"december":12,
    }
    m = re.search(r'(20\d{2})[-_](0[1-9]|1[0-2])', url)
    if m:
        return int(m.group(1)), int(m.group(2))
    m2 = re.search(
        r'(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec)[\s_-]*(20\d{2})',
        url, re.IGNORECASE
    )
    if m2:
        mon = m2.group(1).lower()
        yr = int(m2.group(2))
        return (yr, month_map[mon]) if mon in month_map else None
    return None


def detect_month_from_wb(wb):
    """
    ÙŠÙƒØªØ´Ù Ø§Ù„Ø´Ù‡Ø± Ø§Ù„ÙØ¹Ù„ÙŠ Ù…Ù† Ù…Ø­ØªÙˆÙ‰ Excel:
    ÙŠÙ‚Ø±Ø£ Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ø£ÙŠØ§Ù… ÙÙŠ ØµÙ Ø§Ù„ØªÙˆØ§Ø±ÙŠØ® ÙˆÙŠØ­Ø¯Ø¯ Ø§Ù„Ø´Ù‡Ø±
    Ø¨Ù†Ø§Ø¡Ù‹ Ø¹Ù„Ù‰ Ø£ÙƒØ¨Ø± Ø±Ù‚Ù… ÙŠÙˆÙ… ÙˆØ¹Ø¯Ø¯ Ø£ÙŠØ§Ù… ÙƒÙ„ Ø´Ù‡Ø±.
    """
    import calendar as cal_mod

    now = datetime.now(TZ)

    # Ø¬Ù…Ø¹ Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ø£ÙŠØ§Ù… Ù…Ù† Ø£ÙˆÙ„ sheet Ù…ØªØ§Ø­
    all_day_nums = set()
    for sheet_name, _ in DEPARTMENTS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _, date_row = find_days_and_dates_rows(ws)
        if not date_row:
            continue
        daynum_to_col = get_daynum_to_col(ws, date_row)
        all_day_nums.update(daynum_to_col.keys())
        break

    if not all_day_nums:
        return None

    max_day = max(all_day_nums)
    num_days = len(all_day_nums)
    print(f"  ðŸ” Excel has days 1-{max_day} ({num_days} days total)")

    # Ø¬Ø±Ø¨ Ø§Ù„Ø£Ø´Ù‡Ø±: Ø§Ù„Ø³Ø§Ø¨Ù‚ØŒ Ø§Ù„Ø­Ø§Ù„ÙŠØŒ Ø§Ù„Ù‚Ø§Ø¯Ù…ØŒ Ø¨Ø¹Ø¯ Ø§Ù„Ù‚Ø§Ø¯Ù…
    candidates = [
        add_months(now.year, now.month, -1),
        (now.year, now.month),
        add_months(now.year, now.month, +1),
        add_months(now.year, now.month, +2),
    ]

    for y, m in candidates:
        days_in = cal_mod.monthrange(y, m)[1]
        # Ø§Ù„Ø´Ù‡Ø± Ø§Ù„ØµØ­ÙŠØ­: Ø¹Ø¯Ø¯ Ø£ÙŠØ§Ù…Ù‡ = Ø¹Ø¯Ø¯ Ø§Ù„Ø£ÙŠØ§Ù… ÙÙŠ Excel
        if num_days == days_in:
            print(f"  âœ… Matched: {y}-{m:02d} has {days_in} days")
            return y, m

    # Ø¥Ø°Ø§ Ù„Ù… ÙŠØ·Ø§Ø¨Ù‚ Ø¨Ø§Ù„Ø¶Ø¨Ø·ØŒ Ø®Ø° Ø£ÙˆÙ„ Ø´Ù‡Ø± Ù„Ø§ ÙŠØªØ¬Ø§ÙˆØ² max_day
    for y, m in candidates:
        days_in = cal_mod.monthrange(y, m)[1]
        if max_day <= days_in:
            print(f"  âš ï¸  Best guess: {y}-{m:02d}")
            return y, m

    return None


def main():
    parser = argparse.ArgumentParser(description='Generate employee schedules from roster Excel')
    parser.add_argument('--month', help='ØªØ­Ø¯ÙŠØ¯ Ø§Ù„Ø´Ù‡Ø± ÙŠØ¯ÙˆÙŠØ§Ù‹ YYYY-MM. Ø§ØªØ±ÙƒÙ‡ ÙØ§Ø±ØºØ§Ù‹ Ù„Ù„ÙƒØ´Ù Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ.', default=None)
    args = parser.parse_args()

    if not EXCEL_URL:
        raise RuntimeError("âŒ EXCEL_URL environment variable is missing")

    print("=" * 60)
    print("ðŸ—“ï¸  Employee Schedule Generator")
    print("=" * 60)

    # ØªØ­Ù…ÙŠÙ„ Excel
    data = download_excel(EXCEL_URL)
    wb = load_workbook(BytesIO(data), data_only=True)

    if args.month:
        # Ø´Ù‡Ø± Ù…Ø­Ø¯Ø¯ ÙŠØ¯ÙˆÙŠØ§Ù‹
        try:
            year, month = [int(x) for x in args.month.split('-')]
        except Exception:
            raise RuntimeError('âŒ ØµÙŠØºØ© Ø®Ø§Ø·Ø¦Ø©. Ø§Ø³ØªØ®Ø¯Ù… YYYY-MM Ù…Ø«Ù„ 2026-03')
        print(f"ðŸ“… Month (manual): {year}-{month:02d}")
    else:
        # â”€â”€ Ø§Ù„ÙƒØ´Ù Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # 1) Ù…Ù† Ø±Ø§Ø¨Ø· EXCEL_URL
        detected = detect_month_from_url(EXCEL_URL)
        if detected:
            year, month = detected
            print(f"ðŸ“… Month detected from URL: {year}-{month:02d}")
        else:
            # 2) Ù…Ù† Ù…Ø­ØªÙˆÙ‰ Excel (Ø¹Ø¯Ø¯ Ø§Ù„Ø£ÙŠØ§Ù…)
            print("ðŸ” Detecting month from Excel content...")
            detected = detect_month_from_wb(wb)
            if detected:
                year, month = detected
                print(f"ðŸ“… Month detected from Excel content: {year}-{month:02d}")
            else:
                # 3) Ø¢Ø®Ø± Ø®ÙŠØ§Ø±: Ø§Ù„Ø´Ù‡Ø± Ø§Ù„Ø­Ø§Ù„ÙŠ
                now = datetime.now(TZ)
                year, month = now.year, now.month
                print(f"âš ï¸  Could not detect â€” using current month: {year}-{month:02d}")

    print(f"\n{'=' * 60}")
    print(f"ðŸ“… Processing: {year}-{month:02d}")
    print(f"{'=' * 60}")

    generate_employee_schedules(wb, year, month)
    generate_schedule_index()

    print("\n" + "=" * 60)
    print("âœ… All done!")
    print("=" * 60)
    print(f"\nðŸ“‚ Files saved to: docs/schedules/")
    print(f"ðŸŒ Access at: https://your-site.github.io/roster-site/schedules/")


if __name__ == "__main__":
    main()
