#!/usr/bin/env python3
"""
generate_employee_schedules.py

سكريبت مستقل لتوليد ملفات JSON لجداول الموظفين
يعمل بشكل منفصل عن generate_and_send.py

الاستخدام:
    python generate_employee_schedules.py
    python generate_employee_schedules.py --month 2026-03
"""

import os
import re
import json
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO
from collections import defaultdict

import requests
from openpyxl import load_workbook


# =========================
# Settings
# =========================
EXCEL_URL = os.environ.get("EXCEL_URL", "").strip()
TZ = ZoneInfo("Asia/Muscat")

DEPARTMENTS = [
    ("Officers", "Officers"),
    ("Supervisors", "Supervisors"),
    ("Load Control", "Load Control"),
    ("Export Checker", "Export Checker"),
    ("Export Operators", "Export Operators"),
    ("Unassigned", "Unassigned"),
]

DAYS = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

SHIFT_MAP = {
    "MN06": ("🌅 Morning (MN06)", "Morning"),
    "ME06": ("🌅 Morning (ME06)", "Morning"),
    "ME07": ("🌅 Morning (ME07)", "Morning"),
    "MN12": ("🌆 Afternoon (MN12)", "Afternoon"),
    "AN13": ("🌆 Afternoon (AN13)", "Afternoon"),
    "AE14": ("🌆 Afternoon (AE14)", "Afternoon"),
    "NN21": ("🌙 Night (NN21)", "Night"),
    "NE22": ("🌙 Night (NE22)", "Night"),
}


# =========================
# Helper Functions
# =========================
def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00A0", " ")).strip()


def to_western_digits(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    arabic = {'٠':'0','١':'1','٢':'2','٣':'3','٤':'4','٥':'5','٦':'6','٧':'7','٨':'8','٩':'9'}
    farsi = {'۰':'0','۱':'1','۲':'2','۳':'3','۴':'4','۵':'5','۶':'6','۷':'7','۸':'8','۹':'9'}
    mp = {**arabic, **farsi}
    return "".join(mp.get(ch, ch) for ch in s)


def norm(s) -> str:
    return clean(to_western_digits(s))


def looks_like_time(s: str) -> bool:
    up = norm(s).upper()
    return bool(
        re.match(r"^\d{3,4}\s*H?\s*-\s*\d{3,4}\s*H?$", up)
        or re.match(r"^\d{3,4}\s*H$", up)
        or re.match(r"^\d{3,4}$", up)
    )


def looks_like_employee_name(s: str) -> bool:
    v = norm(s)
    if not v:
        return False
    up = v.upper()
    if looks_like_time(up):
        return False
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", up):
        return False
    if re.search(r"-\s*\d{3,}", v) and re.search(r"[A-Za-z\u0600-\u06FF]", v):
        return True
    parts = [p for p in v.split(" ") if p]
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", v) and len(parts) >= 2)


def looks_like_shift_code(s: str) -> bool:
    v = norm(s).upper()
    if not v:
        return False
    if looks_like_time(v):
        return False
    if v in ["OFF", "O", "LV", "TR", "ST", "SL", "AL", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14", "OT"]:
        return True
    if re.match(r"^(MN|AN|NN|NT|ME|AE|NE)\d{1,2}", v):
        return True
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", v):
        return True
    if len(v) >= 3 and re.search(r"[A-Z]", v):
        return True
    return False


def map_shift(code: str):
    c0 = norm(code)
    c = c0.upper()
    if not c or c == "0":
        return ("-", "Other")
    
    if c == "AL" or c == "LV" or "ANNUAL LEAVE" in c:
        return ("✈️ Annual Leave", "Annual Leave")
    
    if c == "SL" or "SICK LEAVE" in c:
        return ("🤒 Sick Leave", "Sick Leave")
    
    if c in ["TR"] or "TRAINING" in c:
        return ("📚 Training", "Training")
    
    if c in ["ST", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14"] or "STANDBY" in c:
        return (f"🧍 {c0}", "Standby")
    
    if c == "OT" or c.startswith("OT"):
        return (f"⏱️ {c0}", "Standby")
    
    if c in ["OFF", "O"] or re.search(r"(REST|OFF\s*DAY|REST\/OFF)", c):
        return ("🛌 Off Day", "Off Day")
    
    if c in SHIFT_MAP:
        return SHIFT_MAP[c]
    
    return (f"❓ {c0}", "Other")


def download_excel(url: str) -> bytes:
    print(f"📥 Downloading Excel from: {url[:50]}...")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    print("✅ Excel downloaded successfully")
    return r.content


def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]


def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    count = 0
    for d in DAYS:
        if any(d in x for x in ups):
            count += 1
    return count


def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False


def find_days_and_dates_rows(ws, scan_rows: int = 80):
    max_r = min(ws.max_row, scan_rows)
    days_row = None
    
    for r in range(1, max_r + 1):
        vals = _row_values(ws, r)
        if _count_day_tokens(vals) >= 3:
            days_row = r
            break
    
    if not days_row:
        return None, None
    
    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        vals = _row_values(ws, r)
        nums = sum(1 for v in vals if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break
    
    return days_row, date_row


def find_employee_col(ws, start_row: int):
    for c in range(1, min(10, ws.max_column + 1)):
        val = norm(ws.cell(row=start_row, column=c).value)
        if looks_like_employee_name(val):
            return c
    return None


def get_daynum_to_col(ws, date_row: int):
    daynum_to_col = {}
    for c in range(1, ws.max_column + 1):
        val = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(val):
            daynum_to_col[int(float(val))] = c
    return daynum_to_col


def extract_employee_id(name_str):
    """يستخرج الرقم من نص مثل: Ahmed Ali - 12345"""
    match = re.search(r'-\s*(\d+)\s*$', name_str)
    if match:
        return match.group(1).strip()
    return None


# =========================
# Main Functions
# =========================
def generate_employee_schedules(wb, year: int, month: int):
    """
    توليد ملفات JSON لجداول كل موظف
    """
    print(f"\n📅 Generating employee schedules for {year}-{month:02d}...")
    
    all_employees = defaultdict(lambda: {
        "name": "",
        "id": "",
        "department": "",
        "schedules": {}
    })
    
    # معالجة كل قسم
    for sheet_name, dept_name in DEPARTMENTS:
        if sheet_name not in wb.sheetnames:
            continue
        
        print(f"  📋 Processing {dept_name}...")
        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        
        if not (days_row and date_row):
            print(f"    ⚠️  Could not find days/dates rows")
            continue
        
        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        daynum_to_col = get_daynum_to_col(ws, date_row)
        
        if not emp_col:
            print(f"    ⚠️  Could not find employee column")
            continue
        
        emp_count = 0
        # معالجة كل موظف
        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue
            
            emp_id = extract_employee_id(name)
            if not emp_id:
                continue
            
            emp_name = re.sub(r'\s*-\s*\d+\s*$', '', name).strip()
            
            # قراءة مناوبات الشهر
            month_schedule = []
            for day_num in sorted(daynum_to_col.keys()):
                col = daynum_to_col[day_num]
                raw = norm(ws.cell(row=r, column=col).value)
                
                if looks_like_shift_code(raw):
                    label, group = map_shift(raw)
                    
                    try:
                        date_obj = datetime(year, month, day_num, tzinfo=TZ)
                        day_name_ar = ["الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت"]
                        day_name_en = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
                        dow = (date_obj.weekday() + 1) % 7
                        
                        month_schedule.append({
                            "date": date_obj.strftime("%Y-%m-%d"),
                            "day": day_num,
                            "day_name_ar": day_name_ar[dow],
                            "day_name_en": day_name_en[dow],
                            "shift_code": raw.upper(),
                            "shift_label": label,
                            "shift_group": group
                        })
                    except ValueError:
                        continue
            
            if month_schedule:
                month_key = f"{year}-{month:02d}"
                all_employees[emp_id]["name"] = emp_name
                all_employees[emp_id]["id"] = emp_id
                all_employees[emp_id]["department"] = dept_name
                all_employees[emp_id]["schedules"][month_key] = month_schedule
                emp_count += 1
        
        print(f"    ✅ Processed {emp_count} employees")
    
    # حفظ ملفات JSON
    schedules_dir = "docs/schedules"
    os.makedirs(schedules_dir, exist_ok=True)
    
    saved_count = 0
    for emp_id, data in all_employees.items():
        filepath = f"{schedules_dir}/{emp_id}.json"
        
        # قراءة البيانات القديمة
        existing_data = {"name": "", "id": emp_id, "department": "", "schedules": {}}
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except:
                pass
        
        # دمج البيانات
        existing_data["name"] = data["name"]
        existing_data["department"] = data["department"]
        existing_data["schedules"].update(data["schedules"])
        
        # حفظ
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        
        saved_count += 1
    
    print(f"\n✅ Generated schedules for {saved_count} employees")
    return saved_count


def generate_schedule_index():
    """
    ينشئ ملف index لقائمة الموظفين
    """
    print("\n📑 Generating index...")
    schedules_dir = "docs/schedules"
    
    if not os.path.exists(schedules_dir):
        print("  ⚠️  Schedules directory not found")
        return
    
    employees_list = []
    
    for filename in os.listdir(schedules_dir):
        if filename.endswith('.json') and filename != 'index.json':
            emp_id = filename.replace('.json', '')
            filepath = os.path.join(schedules_dir, filename)
            
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    employees_list.append({
                        "id": emp_id,
                        "name": data.get("name", ""),
                        "department": data.get("department", ""),
                        "months": sorted(list(data.get("schedules", {}).keys()))
                    })
            except:
                continue
    
    employees_list.sort(key=lambda x: (x["department"], x["name"]))
    
    index_file = os.path.join(schedules_dir, "index.json")
    with open(index_file, 'w', encoding='utf-8') as f:
        json.dump({
            "total": len(employees_list),
            "employees": employees_list,
            "last_updated": datetime.now(TZ).isoformat()
        }, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Generated index for {len(employees_list)} employees")


# =========================
# Main
# =========================
def add_months(year, month, delta):
    """إضافة/طرح أشهر مع معالجة تجاوز السنة"""
    month += delta
    while month > 12:
        month -= 12
        year += 1
    while month < 1:
        month += 12
        year -= 1
    return year, month


def main():
    parser = argparse.ArgumentParser(description='Generate employee schedules from roster Excel')
    parser.add_argument('--month', help='Month to process (YYYY-MM). Leave empty to process prev+curr+next automatically.', default=None)
    args = parser.parse_args()

    if not EXCEL_URL:
        raise RuntimeError("❌ EXCEL_URL environment variable is missing")

    now = datetime.now(TZ)

    if args.month:
        # شهر محدد يدوياً
        try:
            year, month = [int(x) for x in args.month.split('-')]
            months_to_process = [(year, month)]
        except Exception:
            raise RuntimeError('❌ Invalid month format. Use YYYY-MM (e.g., 2026-03)')
    else:
        # تلقائي: الشهر السابق + الحالي + القادم
        prev_y, prev_m = add_months(now.year, now.month, -1)
        curr_y, curr_m = now.year, now.month
        next_y, next_m = add_months(now.year, now.month, +1)
        months_to_process = [
            (prev_y, prev_m),
            (curr_y, curr_m),
            (next_y, next_m),
        ]

    print("=" * 60)
    print(f"🗓️  Employee Schedule Generator")
    print(f"📅 Months to process: {[f'{y}-{m:02d}' for y, m in months_to_process]}")
    print("=" * 60)

    # تحميل Excel مرة واحدة فقط
    data = download_excel(EXCEL_URL)
    wb = load_workbook(BytesIO(data), data_only=True)

    # توليد الجداول لكل الأشهر
    for year, month in months_to_process:
        print(f"\n{'─'*40}")
        generate_employee_schedules(wb, year, month)

    generate_schedule_index()

    print("\n" + "=" * 60)
    print("✅ All done!")
    print("=" * 60)
    print(f"\n📂 Files saved to: docs/schedules/")
    print(f"🌐 Access at: https://your-site.github.io/roster-site/schedules/")


if __name__ == "__main__":
    main()
