import os
import re
import argparse
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO

import requests
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText


# =========================
# Settings / Secrets
# =========================
EXCEL_URL = os.environ.get("EXCEL_URL", "").strip()

SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "").strip()
SMTP_PASS = os.environ.get("SMTP_PASS", "").strip()
MAIL_FROM = os.environ.get("MAIL_FROM", "").strip()
MAIL_TO = os.environ.get("MAIL_TO", "").strip()

PAGES_BASE_URL = os.environ.get("PAGES_BASE_URL", "").strip()  # optional
TZ = ZoneInfo("Asia/Muscat")
AUTO_OPEN_ACTIVE_SHIFT_IN_FULL = True
# Excel sheets
DEPARTMENTS = [
    ("Officers", "Officers"),
    ("Supervisors", "Supervisors"),
    ("Load Control", "Load Control"),
    ("Export Checker", "Export Checker"),
    ("Export Operators", "Export Operators"),
    ("Unassigned", "Unassigned"),  # â† Ø§Ù„Ù‚Ø³Ù… Ø§Ù„Ø¬Ø¯ÙŠØ¯
]

# For day-row matching only
DAYS = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

SHIFT_MAP = {
    "MN06": ("ğŸŒ… Morning (MN06)", "Morning"),
    "ME06": ("ğŸŒ… Morning (ME06)", "Morning"),
    "ME07": ("ğŸŒ… Morning (ME07)", "Morning"),
    "MN12": ("ğŸŒ† Afternoon (MN12)", "Afternoon"),
    "AN13": ("ğŸŒ† Afternoon (AN13)", "Afternoon"),
    "AE14": ("ğŸŒ† Afternoon (AE14)", "Afternoon"),
    "NN21": ("ğŸŒ™ Night (NN21)", "Night"),
    "NE22": ("ğŸŒ™ Night (NE22)", "Night"),
}

# ØªÙ… ØªØ­ÙˆÙŠÙ„ ÙƒÙ„ Ø§Ù„Ø£Ø³Ù…Ø§Ø¡ Ù„Ù„Ø¥Ù†Ø¬Ù„ÙŠØ²ÙŠØ©
GROUP_ORDER = ["Morning", "Afternoon", "Night", "Standby", "Off Day", "Leave", "Training", "Other"]


# =========================
# Helpers
# =========================
def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00A0", " ")).strip()

def to_western_digits(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    arabic = {'Ù ':'0','Ù¡':'1','Ù¢':'2','Ù£':'3','Ù¤':'4','Ù¥':'5','Ù¦':'6','Ù§':'7','Ù¨':'8','Ù©':'9'}
    farsi  = {'Û°':'0','Û±':'1','Û²':'2','Û³':'3','Û´':'4','Ûµ':'5','Û¶':'6','Û·':'7','Û¸':'8','Û¹':'9'}
    mp = {**arabic, **farsi}
    return "".join(mp.get(ch, ch) for ch in s)

def norm(s) -> str:
    return clean(to_western_digits(s))

def looks_like_time(s: str) -> bool:
    up = norm(s).upper()
    return bool(
        re.match(r"^\d{3,4}\s*H?\s*-\s*\d{3,4}\s*H?$", up)
        or re.match(r"^\d{3,4}\s*H$", up)
        or re.match(r"^\d{3,4}$", up)
    )

def looks_like_employee_name(s: str) -> bool:
    v = norm(s)
    if not v:
        return False
    up = v.upper()
    if looks_like_time(up):
        return False
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", up):
        return False
    # Ù‚ÙˆÙŠ: Ø§Ø³Ù… - Ø±Ù‚Ù…
    if re.search(r"-\s*\d{3,}", v) and re.search(r"[A-Za-z\u0600-\u06FF]", v):
        return True
    # Ø¨Ø¯ÙŠÙ„: ÙƒÙ„Ù…ØªÙŠÙ† Ø£Ùˆ Ø£ÙƒØ«Ø±
    parts = [p for p in v.split(" ") if p]
    return bool(re.search(r"[A-Za-z\u0600-\u06FF]", v) and len(parts) >= 2)

def looks_like_shift_code(s: str) -> bool:
    v = norm(s).upper()
    if not v:
        return False
    if looks_like_time(v):
        return False
    if v in ["OFF", "O", "LV", "TR", "ST", "SL", "AL", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14", "OT"]:
        return True
    if re.match(r"^(MN|AN|NN|NT|ME|AE|NE)\d{1,2}", v):
        return True
    if re.search(r"(ANNUAL\s*LEAVE|SICK\s*LEAVE|REST\/OFF\s*DAY|REST|OFF\s*DAY|TRAINING|STANDBY)", v):
        return True
    # â† Ø¥Ø¶Ø§ÙØ©: Ø£ÙŠ ÙƒÙˆØ¯ ØºØ±ÙŠØ¨ Ù…Ø«Ù„ STAR14 ÙŠØ¹ØªØ¨Ø± shift code
    if len(v) >= 3 and re.search(r"[A-Z]", v):
        return True
    return False

def map_shift(code: str):
    c0 = norm(code)
    c = c0.upper()
    if not c or c == "0":
        return ("-", "Other")

    if c == "AL" or "ANNUAL LEAVE" in c:
        return ("ğŸ–ï¸ Leave", "Leave")
    if c == "SL" or "SICK LEAVE" in c:
        return ("ğŸ¤’ Sick Leave", "Leave")
    if c == "LV":
        return ("ğŸ–ï¸ Leave", "Leave")
    if c in ["TR"] or "TRAINING" in c:
        return ("ğŸ“š Training", "Training")

    # ğŸ”¹ Ø¨Ø§Ù‚ÙŠ Ø§Ù„Ø³ØªØ§Ù†Ø¯Ø¨Ø§ÙŠ
    if c in ["ST", "STM", "STN", "STNE22", "STME06", "STMN06", "STAE14"] or "STANDBY" in c:
        return ("ğŸ§ Standby", "Standby")

    if c == "OT" or c.startswith("OT"):
        return ("â±ï¸ OT", "Standby")
    if c in ["OFF", "O"] or re.search(r"(REST|OFF\s*DAY|REST\/OFF)", c):
        return ("ğŸ›Œ Off Day", "Off Day")

    if c in SHIFT_MAP:
        return SHIFT_MAP[c]

    return (f"â“ {c0}", "Other")

def current_shift_key(now: datetime) -> str:
    # 21:00â€“04:59 Night, 14:00â€“20:59 Afternoon, else Morning
    t = now.hour * 60 + now.minute
    if t >= 21 * 60 or t < 5 * 60:
        return "Night"
    if t >= 14 * 60:
        return "Afternoon"
    return "Morning"

def download_excel(url: str) -> bytes:
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.content

def infer_pages_base_url():
    return "https://khalidsaif912.github.io/roster-site"


# =========================
# Detect rows/cols (Days row + Date numbers row)
# =========================
def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    count = 0
    for d in DAYS:
        if any(d in x for x in ups):
            count += 1
    return count

def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False

def find_days_and_dates_rows(ws, scan_rows: int = 80):
    """
    ÙŠØ¨Ø­Ø« Ø¹Ù† ØµÙ ÙÙŠÙ‡ SUN..SAT Ø¨ÙƒØ«Ø±Ø© Ø«Ù… ØµÙ ØªØ­ØªÙ‡ ÙÙŠÙ‡ Ø£Ø±Ù‚Ø§Ù… 1..31
    """
    max_r = min(ws.max_row, scan_rows)
    days_row = None

    for r in range(1, max_r + 1):
        vals = _row_values(ws, r)
        if _count_day_tokens(vals) >= 3:
            days_row = r
            break

    if not days_row:
        return None, None

    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        vals = _row_values(ws, r)
        nums = sum(1 for v in vals if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break

    return days_row, date_row

def find_day_col(ws, days_row: int, date_row: int, today_dow: int, today_day: int):
    """
    ÙŠØ«Ø¨Øª Ø§Ù„Ø¹Ù…ÙˆØ¯ Ø§Ù„ØµØ­ÙŠØ­ Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„ÙŠÙˆÙ… + Ø±Ù‚Ù… Ø§Ù„ØªØ§Ø±ÙŠØ®
    """
    if not days_row or not date_row:
        return None

    day_key = DAYS[today_dow]
    # Prefer (day + date) match
    for c in range(1, ws.max_column + 1):
        top = norm(ws.cell(row=days_row, column=c).value).upper()
        bot = norm(ws.cell(row=date_row, column=c).value)
        if day_key in top and _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    # Fallback: date-only
    for c in range(1, ws.max_column + 1):
        bot = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    return None


def get_daynum_to_col(ws, date_row: int):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(v):
            m[int(float(v))] = c
    return m

def find_employee_col(ws, start_row: int):
    for c in range(1, min(ws.max_column, 15) + 1):
        found = 0
        for r in range(start_row, min(start_row + 20, ws.max_row) + 1):
            v = norm(ws.cell(row=r, column=c).value)
            if looks_like_employee_name(v):
                found += 1
        if found >= 3:
            return c
    return None

def range_suffix_for_day(day: int, daynum_to_raw: dict, code_key: str):
    """
    Ø¥Ø°Ø§ ÙƒØ§Ù† ÙŠÙˆÙ… (day) Ø¬Ø²Ø¡ Ù…Ù† block Ù…ØªØµÙ„ Ù…Ù† Ù†ÙØ³ code_keyØŒ ÙŠØ±Ø¬Ø¹ (Ù…Ù† X Ø¥Ù„Ù‰ Y)
    """
    sorted_days = sorted(daynum_to_raw.keys())
    if day not in sorted_days:
        return ""

    up_key = code_key.upper()
    start = end = day

    # ØªØ­Ø¯ÙŠØ¯ Ø§Ù„Ø£ÙƒÙˆØ§Ø¯ Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø© Ù„Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹ Ù…Ù† Ø§Ù„Ø¥Ø¬Ø§Ø²Ø©/Ø§Ù„ØªØ¯Ø±ÙŠØ¨
    acceptable_codes = []
    if up_key in ["AL", "LV"] or "ANNUAL" in up_key:
        # Ø§Ù„Ø¥Ø¬Ø§Ø²Ø© Ø§Ù„Ø³Ù†ÙˆÙŠØ©
        acceptable_codes = ["AL", "LV", "ANNUAL LEAVE"]
    elif up_key == "SL" or "SICK" in up_key:
        # Ø§Ù„Ø¥Ø¬Ø§Ø²Ø© Ø§Ù„Ù…Ø±Ø¶ÙŠØ©
        acceptable_codes = ["SL", "SICK LEAVE"]
    elif up_key == "TR" or "TRAINING" in up_key:
        # Ø§Ù„ØªØ¯Ø±ÙŠØ¨
        acceptable_codes = ["TR", "TRAINING"]
    else:
        # Ø£ÙŠ ÙƒÙˆØ¯ Ø¢Ø®Ø± - ÙŠØ¬Ø¨ Ø£Ù† ÙŠÙƒÙˆÙ† Ù…Ø·Ø§Ø¨Ù‚ ØªÙ…Ø§Ù…Ø§Ù‹
        acceptable_codes = [up_key]

    def is_same_type(val: str) -> bool:
        """ØªØ­Ù‚Ù‚ Ø¥Ø°Ø§ ÙƒØ§Ù† Ø§Ù„ÙƒÙˆØ¯ Ù…Ù† Ù†ÙØ³ Ø§Ù„Ù†ÙˆØ¹"""
        val_upper = val.upper()
        for code in acceptable_codes:
            if code in val_upper or val_upper == code:
                return True
        return False

    # backward - Ø§Ù„Ø¨Ø­Ø« Ù„Ù„Ø®Ù„Ù
    for d in reversed(sorted_days):
        if d >= day:
            continue
        val = norm(daynum_to_raw.get(d, ""))
        if is_same_type(val):
            start = d
        else:
            break

    # forward - Ø§Ù„Ø¨Ø­Ø« Ù„Ù„Ø£Ù…Ø§Ù…
    for d in sorted_days:
        if d <= day:
            continue
        val = norm(daynum_to_raw.get(d, ""))
        if is_same_type(val):
            end = d
        else:
            break

    if start == end:
        return ""
    return f"(Ù…Ù† {start} Ø¥Ù„Ù‰ {end})"




# =========================
# Department card colors
# =========================
DEPT_COLORS = [
    {"name": "blue",   "base": "#2563eb", "light": "#2563eb15", "border": "#2563eb18", "grad_from": "#2563eb", "grad_to": "#2563ebcc"},
    {"name": "cyan",   "base": "#0891b2", "light": "#0891b215", "border": "#0891b218", "grad_from": "#0891b2", "grad_to": "#0891b2cc"},
    {"name": "green",  "base": "#059669", "light": "#05966915", "border": "#05966918", "grad_from": "#059669", "grad_to": "#059669cc"},
    {"name": "red",    "base": "#dc2626", "light": "#dc262615", "border": "#dc262618", "grad_from": "#dc2626", "grad_to": "#dc2626cc"},
    {"name": "purple", "base": "#7c3aed", "light": "#7c3aed15", "border": "#7c3aed18", "grad_from": "#7c3aed", "grad_to": "#7c3aedcc"},
    {"name": "orange", "base": "#ea580c", "light": "#ea580c15", "border": "#ea580c18", "grad_from": "#ea580c", "grad_to": "#ea580ccc"},
]

# Ù‚Ø³Ù… Unassigned ÙŠØ£Ø®Ø° Ù„ÙˆÙ† Ø¨Ø±ØªÙ‚Ø§Ù„ÙŠ/Ø±Ù…Ø§Ø¯ÙŠ
UNASSIGNED_COLOR = {"name": "gray", "base": "#6b7280", "light": "#6b728015", "border": "#6b728018", "grad_from": "#6b7280", "grad_to": "#6b7280cc"}

# =========================
# Shift group colors (Morning/Afternoon/Night/etc.)
# =========================
SHIFT_COLORS = {
    "Morning": {
        "border": "#f59e0b44",
        "bg": "#fef3c7",
        "summary_bg": "#fef3c7",
        "summary_border": "#f59e0b33",
        "label_color": "#92400e",
        "count_bg": "#f59e0b22",
        "count_color": "#92400e",
        "status_color": "#92400e",
        "icon": "â˜€ï¸",
    },
    "Afternoon": {
        "border": "#f9731644",
        "bg": "#ffedd5",
        "summary_bg": "#ffedd5",
        "summary_border": "#f9731633",
        "label_color": "#9a3412",
        "count_bg": "#f9731622",
        "count_color": "#9a3412",
        "status_color": "#9a3412",
        "icon": "ğŸŒ¤ï¸",
    },
    "Night": {
        "border": "#8b5cf644",
        "bg": "#ede9fe",
        "summary_bg": "#ede9fe",
        "summary_border": "#8b5cf633",
        "label_color": "#5b21b6",
        "count_bg": "#8b5cf622",
        "count_color": "#5b21b6",
        "status_color": "#5b21b6",
        "icon": "ğŸŒ™",
    },
    "Off Day": {
        "border": "#6366f144",
        "bg": "#e0e7ff",
        "summary_bg": "#e0e7ff",
        "summary_border": "#6366f133",
        "label_color": "#3730a3",
        "count_bg": "#6366f122",
        "count_color": "#3730a3",
        "status_color": "#3730a3",
        "icon": "ğŸ›‹ï¸",
    },
    "Leave": {
        "border": "#10b98144",
        "bg": "#d1fae5",
        "summary_bg": "#d1fae5",
        "summary_border": "#10b98133",
        "label_color": "#065f46",
        "count_bg": "#10b98122",
        "count_color": "#065f46",
        "status_color": "#065f46",
        "icon": "âœˆï¸",
    },
    "Training": {
        "border": "#0ea5e944",
        "bg": "#e0f2fe",
        "summary_bg": "#e0f2fe",
        "summary_border": "#0ea5e933",
        "label_color": "#075985",
        "count_bg": "#0ea5e922",
        "count_color": "#075985",
        "status_color": "#075985",
        "icon": "ğŸ“š",
    },
    "Standby": {
        "border": "#9e9e9e44",
        "bg": "#f0f0f0",
        "summary_bg": "#f0f0f0",
        "summary_border": "#9e9e9e33",
        "label_color": "#555555",
        "count_bg": "#cccccc22",
        "count_color": "#555555",
        "status_color": "#555555",
        "icon": "ğŸ§"
    }, 
    "Other": {
        "border": "#94a3b844",
        "bg": "#f1f5f9",
        "summary_bg": "#f1f5f9",
        "summary_border": "#94a3b833",
        "label_color": "#475569",
        "count_bg": "#94a3b822",
        "count_color": "#475569",
        "status_color": "#475569",
        "icon": "â“",
    },
}


# =========================
# HTML Builders
# =========================
def dept_card_html(dept_name: str, dept_color: dict, buckets: dict, open_group: str = None) -> str:
    # buckets = {group_key: [{"name": ..., "shift": ...}, ...], ...}
    total = sum(len(buckets.get(k, [])) for k in GROUP_ORDER)
    if total == 0:
        return ""

    shifts_html = ""
    for group_key in GROUP_ORDER:
        emps = buckets.get(group_key, [])
        if not emps:
            continue

        # Determine shift display name (use English directly)
        if group_key == "Morning":
            display_name = "Morning"
        elif group_key == "Afternoon":
            display_name = "Afternoon"
        elif group_key == "Night":
            display_name = "Night"
        elif group_key == "Off Day":
            display_name = "Off Day"
        elif group_key == "Leave":
            display_name = "Annual Leave"
        elif group_key == "Training":
            display_name = "Training"
        elif group_key == "Standby":
            display_name = "Standby"
        else:
            display_name = "Other"

        colors = SHIFT_COLORS.get(group_key, SHIFT_COLORS["Other"])
        count = len(emps)
        open_attr = ' open' if (group_key == open_group) else ''

        rows_html = ""
        for i, e in enumerate(emps):
            alt = " empRowAlt" if i % 2 == 1 else ""
            rows_html += f"""<div class="empRow{alt}">
       <span class="empName">{e['name']}</span>
       <span class="empStatus" style="color:{colors['status_color']};">{e['shift']}</span>
     </div>"""

        shifts_html += f"""
    <details class="shiftCard" style="border:1px solid {colors['border']}; background:{colors['bg']}"{open_attr}>
      <summary class="shiftSummary" style="background:{colors['summary_bg']}; border-bottom:1px solid {colors['summary_border']};">
        <span class="shiftIcon">{colors['icon']}</span>
        <span class="shiftLabel" style="color:{colors['label_color']};">{display_name}</span>
        <span class="shiftCount" style="background:{colors['count_bg']}; color:{colors['count_color']};">{count}</span>
      </summary>
      <div class="shiftBody">
        {rows_html}
      </div>
    </details>
            """

    icon_svg = """
<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>
"""

    return f"""
    <div class="deptCard">
      <div style="height:5px; background:linear-gradient(to right, {dept_color['grad_from']}, {dept_color['grad_to']});"></div>

      <div class="deptHead" style="border-bottom:2px solid {dept_color['border']};">
        <div class="deptIcon" style="background:{dept_color['light']}; color:{dept_color['base']};">
          {icon_svg}
        </div>
        <div class="deptTitle">{dept_name}</div>
        <div class="deptBadge" style="background:{dept_color['light']}; color:{dept_color['base']}; border:1px solid {dept_color['border']};">
          <span style="font-size:10px;opacity:.7;display:block;margin-bottom:1px;text-transform:uppercase;letter-spacing:.5px;">Total</span>
          <span style="font-size:17px;font-weight:900;">{total}</span>
        </div>
      </div>

      <div class="shiftStack">
{shifts_html}
      </div>
    </div>
    """


def page_shell_html(date_label: str, iso_date: str, employees_total: int, departments_total: int,
                     dept_cards_html: str, cta_url: str, sent_time: str) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster</title>
  <style>

    /* â•â•â•â•â•â•â• RESET â•â•â•â•â•â•â• */
    body {{
      margin:0; padding:0;
      background:#eef1f7;
      font-family:'Segoe UI', system-ui, -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif;
      color:#0f172a;
      -webkit-font-smoothing:antialiased;
    }}
    * {{ box-sizing:border-box; }}

    /* â•â•â•â•â•â•â• WRAP â•â•â•â•â•â•â• */
    .wrap {{ max-width:680px; margin:0 auto; padding:16px 14px 28px; }}

    /* â•â•â•â•â•â•â• HEADER â•â•â•â•â•â•â• */
    .header {{
      background:linear-gradient(135deg, #1e40af 0%, #1976d2 50%, #0ea5e9 100%);
      color:#fff;
      padding:26px 18px 24px;
      border-radius:20px;
      text-align:center;
      box-shadow:0 8px 28px rgba(30,64,175,.25);
      position:relative;
      overflow:hidden;
    }}
    .header::before {{
      content:''; position:absolute;
      top:-30px; right:-40px;
      width:140px; height:140px;
      border-radius:50%;
      background:rgba(255,255,255,.08);
    }}
    .header::after {{
      content:''; position:absolute;
      bottom:-50px; left:-30px;
      width:160px; height:160px;
      border-radius:50%;
      background:rgba(255,255,255,.06);
    }}
    .header h1 {{ margin:0; font-size:24px; font-weight:800; position:relative; z-index:1; letter-spacing:-.3px; }}
    .header .dateTag {{
      display:inline-block; margin-top:10px;
      background:rgba(255,255,255,.18);
      padding:5px 18px; border-radius:30px;
      font-size:13px; font-weight:600; letter-spacing:.3px;
      position:relative; z-index:1;
    }}

    /* â•â•â•â•â•â•â• SUMMARY BAR â•â•â•â•â•â•â• */
    .summaryBar {{ display:flex; justify-content:center; gap:12px; margin-top:14px; }}
    .summaryChip {{
      background:#fff;
      border:1px solid rgba(15,23,42,.1);
      border-radius:14px;
      padding:10px 20px;
      text-align:center;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
    }}
    .summaryChip .chipVal {{ font-size:22px; font-weight:900; color:#1e40af; }}
    .summaryChip .chipLabel {{ font-size:11px; font-weight:600; color:#64748b; text-transform:uppercase; letter-spacing:.6px; margin-top:2px; }}

    /* â•â•â•â•â•â•â• DEPARTMENT CARD â•â•â•â•â•â•â• */
    .deptCard {{
      margin-top:18px;
      background:#fff;
      border-radius:18px;
      overflow:hidden;
      border:1px solid rgba(15,23,42,.07);
      box-shadow:0 4px 18px rgba(15,23,42,.08);
    }}
    .deptHead {{
      display:flex;
      align-items:center;
      gap:12px;
      padding:14px 16px;
      background:#fff;
    }}
    .deptIcon {{
      width:40px; height:40px;
      border-radius:12px;
      display:flex; align-items:center; justify-content:center;
      flex-shrink:0;
    }}
    .deptTitle {{ font-size:18px; font-weight:800; color:#1e293b; flex:1; letter-spacing:-.2px; }}
    .deptBadge {{ min-width:48px; padding:6px 10px; border-radius:12px; text-align:center; }}

    /* â•â•â•â•â•â•â• SHIFT STACK â•â•â•â•â•â•â• */
    .shiftStack {{ padding:10px; display:flex; flex-direction:column; gap:8px; }}

    /* â•â•â•â•â•â•â• SHIFT CARD â€” <details> â•â•â•â•â•â•â• */
    .shiftCard {{
      border-radius:14px;
      overflow:hidden;
    }}

    .shiftSummary {{
      display:flex;
      align-items:center;
      gap:10px;
      padding:11px 14px;
      cursor:pointer;
      list-style:none;
      -webkit-appearance:none;
      appearance:none;
      user-select:none;
    }}
    .shiftSummary::-webkit-details-marker {{ display:none; }}
    .shiftSummary::marker              {{ display:none; }}

    .shiftIcon  {{ font-size:20px; line-height:1; flex-shrink:0; }}
    .shiftLabel {{ font-size:15px; font-weight:800; flex:1; letter-spacing:-.1px; }}
    .shiftCount {{
      font-size:13px; font-weight:800;
      padding:3px 10px; border-radius:20px;
      flex-shrink:0;
    }}

    /* chevron ÙŠØ¯ÙˆØ± Ù„Ù…Ø§ ÙŠÙØªØ­ */
    .shiftSummary::after {{
      content:'â–¾';
      font-size:14px;
      color:#94a3b8;
      transition:transform .2s;
      flex-shrink:0;
    }}
    .shiftCard[open] .shiftSummary::after {{
      transform:rotate(180deg);
    }}

    .shiftBody {{ background:rgba(255,255,255,.7); }}

    /* â”€â”€ employee row â”€â”€ */
    .empRow {{
      display:flex;
      align-items:center;
      justify-content:space-between;
      padding:9px 16px;
      border-top:1px solid rgba(15,23,42,.06);
    }}
    .empRowAlt {{ background:rgba(15,23,42,.02); }}
    .empName  {{ font-size:15px; font-weight:700; color:#1e293b; }}
    .empStatus {{ font-size:13px; font-weight:600; }}

    /* â•â•â•â•â•â•â• CTA â•â•â•â•â•â•â• */
    .btnWrap {{ margin-top:20px; text-align:center; }}
    .btn {{
      display:inline-block;
      padding:14px 38px;
      border-radius:16px;
      background:linear-gradient(135deg, #1e40af, #1976d2);
      color:#fff !important;
      text-decoration:none;
      font-weight:800;
      font-size:15px;
      box-shadow:0 6px 20px rgba(30,64,175,.3);
    }}

    /* â•â•â•â•â•â•â• FOOTER â•â•â•â•â•â•â• */
    .footer {{ margin-top:18px; text-align:center; font-size:12px; color:#94a3b8; padding:12px 0; line-height:1.9; }}
    .footer strong {{ color:#64748b; }}

    /* â•â•â•â•â•â•â• MOBILE â•â•â•â•â•â•â• */
    @media (max-width:480px){{
      .wrap            {{ padding:12px 10px 22px; }}
      .header h1       {{ font-size:21px; }}
      .deptTitle       {{ font-size:16px; }}
      .empName         {{ font-size:14px; }}
      .empStatus       {{ font-size:12px; }}
      .shiftLabel      {{ font-size:14px; }}
      .summaryBar      {{ gap:8px; }}
      .summaryChip     {{ padding:8px 14px; }}
      .summaryChip .chipVal {{ font-size:19px; }}
    }}

  </style>
</head>
<body>
<div class="wrap">

  <!-- â•â•â•â• HEADER â•â•â•â• -->
  <div class="header">
    <h1>ğŸ“‹ Duty Roster</h1>
    <div class="dateTag" id="dateTag" role="button" tabindex="0" style="cursor:pointer;">ğŸ“… {date_label}</div>
    <input id="datePicker" type="date" value="{iso_date}" style="position:absolute;left:0;top:0;width:1px;height:1px;opacity:0;pointer-events:none;" aria-hidden="true" />
  </div>

  <!-- â•â•â•â• SUMMARY CHIPS â•â•â•â• -->
  <div class="summaryBar">
    <div class="summaryChip">
      <div class="chipVal">{employees_total}</div>
      <div class="chipLabel">Employees</div>
    </div>
    <div class="summaryChip">
      <div class="chipVal" style="color:#059669;">{departments_total}</div>
      <div class="chipLabel">Departments</div>
    </div>
  </div>

  <!-- â•â•â•â• DEPARTMENT CARDS â•â•â•â• -->
  {dept_cards_html}

  <!-- â•â•â•â• CTA â•â•â•â• -->
  <div class="btnWrap">
    <a class="btn" href="{cta_url}">ğŸ“‹ View Full Duty Roster</a>
  </div>

  <!-- â•â•â•â• FOOTER â•â•â•â• -->
  <div class="footer">
    Sent at <strong>{sent_time}</strong>
     &nbsp;Â·&nbsp; Total: <strong>{employees_total} employees</strong>
  </div>

</div>

<script>
(function(){{
  var tag = document.getElementById('dateTag');
  var picker = document.getElementById('datePicker');
  if(!tag || !picker) return;

  function openPicker(){{
    // Position the (hidden) input Ù…Ø¨Ø§Ø´Ø±Ø© ØªØ­Øª Ø§Ù„ØªØ§Ø±ÙŠØ® Ø­ØªÙ‰ ÙŠØ¸Ù‡Ø± Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø¨Ø§Ù„Ù…ÙƒØ§Ù† Ø§Ù„ØµØ­ÙŠØ­
    try{{
      var r = tag.getBoundingClientRect();
      var wrap = tag.closest('.header') || document.body;
      var wr = wrap.getBoundingClientRect();
      picker.style.left = (r.left - wr.left) + 'px';
      picker.style.top  = (r.bottom - wr.top + 6) + 'px';
      picker.style.width = Math.max(120, r.width) + 'px';
    }}catch(e){{}}

    try{{
      if (picker.showPicker) {{ picker.showPicker(); }}
      else {{ picker.focus(); picker.click(); }}
    }}catch(e){{
      picker.focus(); picker.click();
    }}
  }}

  tag.addEventListener('click', openPicker);
  tag.addEventListener('keydown', function(e){{
    if(e.key === 'Enter' || e.key === ' ') {{ e.preventDefault(); openPicker(); }}
  }});

  function computeBasePath(){{
    var p = window.location.pathname || '/';
    // Strip "/now/" and anything after it
    p = p.replace(/\/now\/.*$/,'/');
    // Strip "/date/YYYY-MM-DD/" and anything after it
    p = p.replace(/\/date\/\\d{{4}}-\\d{{2}}-\\d{{2}}\\/.*$/,'/');
    return p.replace(/\/+$/,''); // no trailing slash
  }}

  picker.addEventListener('change', function(){{
    if(!picker.value) return;
    var base = computeBasePath();
    var target = base + '/date/' + picker.value + '/';
    window.location.href = target;
  }});
}})();
</script>

</body>
</html>"""


def generate_date_pages_for_month(wb, year: int, month: int, pages_base: str):
    """
    Generate static pages for each day of the given month.
    Used by the date picker to navigate to different dates.
    """
    import calendar
    from datetime import datetime as dt

    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        try:
            date_obj = dt(year, month, day, tzinfo=TZ)
            dow = (date_obj.weekday() + 1) % 7  # Sun=0
            active_group = current_shift_key(date_obj)

            dept_cards_all = []
            dept_cards_now = []
            employees_total_all = 0
            employees_total_now = 0
            depts_count = 0

            for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                days_row, date_row = find_days_and_dates_rows(ws)
                day_col = find_day_col(ws, days_row, date_row, dow, day)

                if not (days_row and date_row and day_col):
                    continue

                start_row = date_row + 1
                emp_col = find_employee_col(ws, start_row=start_row)
                daynum_to_col = get_daynum_to_col(ws, date_row)
                if not emp_col:
                    continue

                buckets = {k: [] for k in GROUP_ORDER}
                buckets_now = {k: [] for k in GROUP_ORDER}

                for r in range(start_row, ws.max_row + 1):
                    name = norm(ws.cell(row=r, column=emp_col).value)
                    if not looks_like_employee_name(name):
                        continue

                    daynum_to_raw = {dn: norm(ws.cell(row=r, column=col).value) for dn, col in daynum_to_col.items()}
                    raw = daynum_to_raw.get(day, "")
                    if not looks_like_shift_code(raw):
                        continue

                    label, grp = map_shift(raw)

                    up = norm(raw).upper()
                    if grp == "Leave":
                        if up == "AL" or "ANNUAL LEAVE" in up or up == "LV":
                            suf = range_suffix_for_day(day, daynum_to_raw, "AL")
                            if suf:
                                label = f"{label} {suf}"
                        elif up == "SL" or "SICK LEAVE" in up:
                            suf = range_suffix_for_day(day, daynum_to_raw, "SL")
                            if suf:
                                label = f"{label} {suf}"
                    elif grp == "Training":
                        if up == "TR" or "TRAINING" in up:
                            suf = range_suffix_for_day(day, daynum_to_raw, "TR")
                            if suf:
                                label = f"{label} {suf}"

                    buckets.setdefault(grp, []).append({"name": name, "shift": label})

                    if grp == active_group:
                        buckets_now.setdefault(grp, []).append({"name": name, "shift": label})

                # ØªØ­Ø¯ÙŠØ¯ Ø§Ù„Ù„ÙˆÙ† Ù„Ù„Ù‚Ø³Ù…
                if dept_name == "Unassigned":
                    dept_color = UNASSIGNED_COLOR
                else:
                    dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]

                open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None
                card_all = dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full)
                dept_cards_all.append(card_all)

                card_now = dept_card_html(dept_name, dept_color, buckets_now, open_group=active_group)
                dept_cards_now.append(card_now)

                employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
                employees_total_now += sum(len(buckets_now.get(g, [])) for g in GROUP_ORDER)

                depts_count += 1

            date_label = date_obj.strftime("%-d %B %Y") if hasattr(date_obj, "strftime") else date_obj.strftime("%d %B %Y")
            try:
                date_label = date_obj.strftime("%-d %B %Y")
            except Exception:
                date_label = date_obj.strftime("%d %B %Y")

            iso_date = date_obj.strftime("%Y-%m-%d")
            sent_time = date_obj.strftime("%H:%M")

            full_url = f"{pages_base}/"
            now_url = f"{pages_base}/now/"

            html_full = page_shell_html(
                date_label=date_label,
                iso_date=iso_date,
                employees_total=employees_total_all,
                departments_total=depts_count,
                dept_cards_html="\n".join(dept_cards_all),
                cta_url=now_url,
                sent_time=sent_time,
            )
            html_now = page_shell_html(
                date_label=date_label,
                iso_date=iso_date,
                employees_total=employees_total_now,
                departments_total=depts_count,
                dept_cards_html="\n".join(dept_cards_now),
                cta_url=full_url,
                sent_time=sent_time,
            )

            date_dir = f"docs/date/{iso_date}"
            os.makedirs(date_dir, exist_ok=True)

            with open(f"{date_dir}/index.html", "w", encoding="utf-8") as f:
                f.write(html_full)

        except Exception as e:
            print(f"Skipping {year}-{month:02d}-{day:02d}: {e}")
            continue


def build_pretty_email_html(active_shift_key: str, now: datetime, all_shifts_by_dept: list, pages_base: str) -> str:
    """
    Builds a beautifully formatted HTML email showing ALL shifts for today.
    all_shifts_by_dept = [{"dept": ..., "shifts": {"Morning": [...], "Afternoon": [...], ...}}, ...]
    """
    # Calculate totals across all shifts
    total_employees = 0
    depts_with_employees = 0
    
    for d in all_shifts_by_dept:
        dept_total = sum(len(employees) for employees in d["shifts"].values())
        if dept_total > 0:
            depts_with_employees += 1
            total_employees += dept_total

    # Determine current shift colors for header
    shift_colors = SHIFT_COLORS.get(active_shift_key, SHIFT_COLORS["Other"])
    shift_icon = shift_colors.get("icon", "â°")

    # Build department cards with ALL shifts
    dept_cards = []
    for idx, d in enumerate(all_shifts_by_dept):
        dept_name = d["dept"]
        shifts_data = d["shifts"]
        
        # Skip if department has no employees today
        dept_total = sum(len(employees) for employees in shifts_data.values())
        if dept_total == 0:
            continue
        
        # Determine department color
        if dept_name == "Unassigned":
            dept_color = UNASSIGNED_COLOR
        else:
            dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]

        # Build shift sections
        shift_sections = ""
        for group_key in GROUP_ORDER:
            employees = shifts_data.get(group_key, [])
            if not employees:
                continue

            # Get shift display name
            shift_display_names = {
                "Morning": "Morning",
                "Afternoon": "Afternoon",
                "Night": "Night",
                "Off Day": "Off Day",
                "Leave": "Annual Leave",
                "Training": "Training",
                "Standby": "Standby",
                "Other": "Other"
            }
            display_name = shift_display_names.get(group_key, group_key)
            
            colors = SHIFT_COLORS.get(group_key, SHIFT_COLORS["Other"])
            count = len(employees)

            # Highlight active shift
            is_active = (group_key == active_shift_key)
            active_border = f"border:2px solid {colors['border']};" if is_active else f"border:1px solid {colors['border']};"
            active_badge = "âš¡" if is_active else ""

            # Build employee rows
            rows_html = ""
            for i, e in enumerate(employees):
                bg_color = "rgba(15,23,42,.03)" if i % 2 == 1 else "transparent"
                rows_html += f"""
                    <tr>
                      <td style="padding:10px 14px;border-top:1px solid rgba(15,23,42,.06);background:{bg_color};">
                        <span style="font-size:14px;font-weight:700;color:#1e293b;">{e['name']}</span>
                      </td>
                      <td style="padding:10px 14px;border-top:1px solid rgba(15,23,42,.06);text-align:right;background:{bg_color};">
                        <span style="font-size:13px;font-weight:600;color:{colors['status_color']};white-space:nowrap;">{e['shift']}</span>
                      </td>
                    </tr>"""

            shift_sections += f"""
              <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:10px;background:{colors['bg']};border-radius:12px;overflow:hidden;{active_border}">
                <!-- Shift Header -->
                <tr>
                  <td colspan="2" style="padding:10px 14px;background:{colors['summary_bg']};border-bottom:1px solid {colors['summary_border']};">
                    <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                      <tr>
                        <td style="padding:0;">
                          <span style="font-size:18px;margin-right:8px;">{colors['icon']}</span>
                          <span style="font-size:15px;font-weight:800;color:{colors['label_color']};letter-spacing:-.1px;">{display_name} {active_badge}</span>
                        </td>
                        <td style="text-align:right;padding:0;">
                          <span style="display:inline-block;padding:4px 12px;border-radius:20px;background:{colors['count_bg']};color:{colors['count_color']};font-size:13px;font-weight:800;">{count}</span>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
                <!-- Employees -->
                {rows_html}
              </table>"""

        # Department icon SVG
        icon_svg = """<svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>"""

        dept_cards.append(f"""
          <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:18px;background:#fff;border-radius:18px;overflow:hidden;border:1px solid rgba(15,23,42,.07);box-shadow:0 4px 18px rgba(15,23,42,.08);">
            <!-- Colored top gradient bar -->
            <tr>
              <td colspan="2" style="height:5px;background:linear-gradient(to right,{dept_color['grad_from']},{dept_color['grad_to']});padding:0;"></td>
            </tr>
            
            <!-- Department Header -->
            <tr>
              <td colspan="2" style="padding:14px 16px;border-bottom:2px solid {dept_color['border']};background:#fff;">
                <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                  <tr>
                    <td style="width:46px;padding:0;">
                      <div style="width:44px;height:44px;border-radius:12px;background:{dept_color['light']};color:{dept_color['base']};display:flex;align-items:center;justify-content:center;">
                        {icon_svg}
                      </div>
                    </td>
                    <td style="padding:0 0 0 12px;">
                      <span style="font-size:18px;font-weight:800;color:#1e293b;letter-spacing:-.2px;display:block;">{dept_name}</span>
                    </td>
                    <td style="text-align:right;padding:0;">
                      <div style="display:inline-block;min-width:52px;padding:8px 12px;border-radius:12px;background:{dept_color['light']};border:1px solid {dept_color['border']};text-align:center;">
                        <span style="font-size:10px;opacity:.7;display:block;text-transform:uppercase;letter-spacing:.5px;color:{dept_color['base']};margin-bottom:1px;">Total</span>
                        <span style="font-size:17px;font-weight:900;color:{dept_color['base']};display:block;">{dept_total}</span>
                      </div>
                    </td>
                  </tr>
                </table>
              </td>
            </tr>

            <!-- All Shifts -->
            <tr>
              <td colspan="2" style="padding:10px;">
                {shift_sections}
              </td>
            </tr>
          </table>
        """)

    dept_html = "".join(dept_cards)
    sent_time = now.strftime("%H:%M")
    date_str = now.strftime("%d %B %Y")

    # Translate active_shift_key display
    shift_display_map = {
        "Morning": "Morning Shift",
        "Afternoon": "Afternoon Shift", 
        "Night": "Night Shift"
    }
    shift_display = shift_display_map.get(active_shift_key, active_shift_key)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster - {date_str}</title>
  <style>
    @media only screen and (max-width: 600px) {{
      .mobile-padding {{ padding: 12px !important; }}
      .mobile-font {{ font-size: 13px !important; }}
      .header-icon {{ font-size: 56px !important; }}
    }}
  </style>
</head>
<body style="margin:0;padding:0;background:#eef1f7;font-family:'Segoe UI',system-ui,-apple-system,BlinkMacSystemFont,Roboto,Helvetica,Arial,sans-serif;-webkit-font-smoothing:antialiased;">
  
  <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#eef1f7;">
    <tr>
      <td align="center" style="padding:20px 14px;">
        
        <!-- Main Container -->
        <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="max-width:680px;width:100%;margin:0 auto;">
          
          <!-- Large Header with Gradient -->
          <tr>
            <td style="padding:0;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:linear-gradient(135deg, #1e40af 0%, #1976d2 50%, #0ea5e9 100%);border-radius:20px 20px 0 0;overflow:hidden;box-shadow:0 10px 35px rgba(30,64,175,.3);position:relative;">
                <tr>
                  <td style="padding:22px 20px;text-align:center;position:relative;">
                    <!-- Decorative circles -->
                    <div style="position:absolute;top:-40px;right:-50px;width:180px;height:180px;border-radius:50%;background:rgba(255,255,255,.08);"></div>
                    <div style="position:absolute;bottom:-60px;left:-40px;width:200px;height:200px;border-radius:50%;background:rgba(255,255,255,.06);"></div>
                    
                    <!-- Large Icon -->
                    <div class="header-icon" style="font-size:44px;margin-bottom:10px;">ğŸ“‹</div>
                    
                    <h1 style="margin:0;font-size:24px;font-weight:800;color:#ffffff;letter-spacing:-.5px;position:relative;z-index:1;">Duty Roster</h1>
                    
                    <div style="margin-top:16px;display:inline-block;background:rgba(255,255,255,.22);padding:10px 26px;border-radius:30px;font-size:16px;font-weight:700;color:#ffffff;letter-spacing:.4px;position:relative;z-index:1;">
                      {shift_icon} {shift_display} <span style="opacity:.8;">(Active Now)</span>
                    </div>
                    
                    <div style="margin-top:10px;display:inline-block;background:rgba(255,255,255,.16);padding:8px 24px;border-radius:30px;font-size:15px;font-weight:600;color:#ffffff;letter-spacing:.3px;position:relative;z-index:1;">
                      ğŸ“… {date_str}
                    </div>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Summary Stats -->
          <tr>
            <td style="padding:0 14px;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;margin-top:18px;">
                <tr>
                  <td style="width:50%;padding-right:6px;">
                    <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border:1px solid rgba(15,23,42,.10);border-radius:16px;box-shadow:0 3px 12px rgba(15,23,42,.07);">
                      <tr>
                        <td style="padding:16px;text-align:center;">
                          <div style="font-size:28px;font-weight:900;color:#1e40af;margin-bottom:4px;">{total_employees}</div>
                          <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;">Employees</div>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <td style="width:50%;padding-left:6px;">
                    <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border:1px solid rgba(15,23,42,.10);border-radius:16px;box-shadow:0 3px 12px rgba(15,23,42,.07);">
                      <tr>
                        <td style="padding:16px;text-align:center;">
                          <div style="font-size:28px;font-weight:900;color:#059669;margin-bottom:4px;">{depts_with_employees}</div>
                          <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;">Departments</div>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Department Cards with ALL Shifts -->
          <tr>
            <td style="padding:0 14px;">
              {dept_html}
            </td>
          </tr>

          <!-- Call to Action Buttons -->
          <tr>
            <td style="padding:22px 14px;text-align:center;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
                <tr>
                  <td style="padding:0 7px 0 0;">
                    <a href="{pages_base}/now/" style="display:inline-block;padding:15px 30px;border-radius:16px;background:linear-gradient(135deg,#1e40af,#1976d2);color:#ffffff;text-decoration:none;font-weight:800;font-size:15px;box-shadow:0 6px 22px rgba(30,64,175,.35);white-space:nowrap;">
                      ğŸ”„ Refresh Now
                    </a>
                  </td>
                  <td style="padding:0 0 0 7px;">
                    <a href="{pages_base}/" style="display:inline-block;padding:15px 30px;border-radius:16px;background:linear-gradient(135deg,#0ea5e9,#06b6d4);color:#ffffff;text-decoration:none;font-weight:800;font-size:15px;box-shadow:0 6px 22px rgba(14,165,233,.35);white-space:nowrap;">
                      ğŸ“‹ View Full Roster
                    </a>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="padding:0 14px 22px;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border-radius:0 0 20px 20px;border:1px solid rgba(15,23,42,.08);border-top:none;">
                <tr>
                  <td style="padding:18px;text-align:center;color:#94a3b8;font-size:13px;line-height:1.9;">
                    Sent at <strong style="color:#64748b;">{sent_time}</strong>
                    <br>
                    Total on duty: <strong style="color:#64748b;">{total_employees} employees</strong> across <strong style="color:#64748b;">{depts_with_employees} departments</strong>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

        </table>
        
      </td>
    </tr>
  </table>

</body>
</html>"""




# =========================
# Email
# =========================
def send_email(subject: str, html: str):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASS and MAIL_FROM and MAIL_TO):
        return
    msg = MIMEText(html, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = MAIL_FROM
    msg["To"] = MAIL_TO

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.sendmail(MAIL_FROM, [x.strip() for x in MAIL_TO.split(",") if x.strip()], msg.as_string())


# =========================
# Main
# =========================
def main():
    if not EXCEL_URL:
        raise RuntimeError("EXCEL_URL missing")

    parser = argparse.ArgumentParser(description='Generate roster pages and send email')
    parser.add_argument('--date', help='Override roster date (YYYY-MM-DD)')
    args = parser.parse_args()

    now = datetime.now(TZ)
    if args.date:
        try:
            y, m, d = [int(x) for x in args.date.strip().split('-')]
            now = datetime(y, m, d, now.hour, now.minute, tzinfo=TZ)
        except Exception:
            raise RuntimeError('Invalid --date format. Use YYYY-MM-DD')
    
    today_dow = (now.weekday() + 1) % 7
    today_day = now.day

    active_group = current_shift_key(now)
    
    # â† Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ pages_base ÙˆØªÙ†Ø¸ÙŠÙÙ‡
    pages_base_raw = PAGES_BASE_URL or infer_pages_base_url()
    # Ø¥Ø²Ø§Ù„Ø© Ø£ÙŠ Ø³Ù„Ø§Ø´ Ù…Ù† Ø§Ù„Ù†Ù‡Ø§ÙŠØ©
    pages_base = pages_base_raw.rstrip("/")
    # Ø¥Ø²Ø§Ù„Ø© /now Ù…Ù† Ø§Ù„Ù†Ù‡Ø§ÙŠØ© Ø¥Ù† ÙˆØ¬Ø¯Øª
    if pages_base.endswith("/now"):
        pages_base = pages_base[:-4]
    
    data = download_excel(EXCEL_URL)
    wb = load_workbook(BytesIO(data), data_only=True)

    # Generate static pages for each date in the current month (used by the date picker)
    generate_date_pages_for_month(wb, now.year, now.month, pages_base)

    dept_cards_all = []
    dept_cards_now = []
    all_shifts_by_dept = []
    employees_total_all = 0
    employees_total_now = 0
    depts_count = 0

    for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        day_col = find_day_col(ws, days_row, date_row, today_dow, today_day)

        if not (days_row and date_row and day_col):
            continue

        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        daynum_to_col = get_daynum_to_col(ws, date_row)
        if not emp_col:
            continue

        buckets = {k: [] for k in GROUP_ORDER}
        buckets_now = {k: [] for k in GROUP_ORDER}

        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue

            daynum_to_raw = {dn: norm(ws.cell(row=r, column=col).value) for dn, col in daynum_to_col.items()}

            raw = daynum_to_raw.get(today_day, "")
            if not looks_like_shift_code(raw):
                continue

            label, grp = map_shift(raw)

            up = norm(raw).upper()
            if grp == "Leave":
                if up == "AL" or "ANNUAL LEAVE" in up or up == "LV":
                    suf = range_suffix_for_day(today_day, daynum_to_raw, "AL")
                    if suf:
                        label = f"{label} {suf}"
                elif up == "SL" or "SICK LEAVE" in up:
                    suf = range_suffix_for_day(today_day, daynum_to_raw, "SL")
                    if suf:
                        label = f"{label} {suf}"
            elif grp == "Training":
                if up == "TR" or "TRAINING" in up:
                    suf = range_suffix_for_day(today_day, daynum_to_raw, "TR")
                    if suf:
                        label = f"{label} {suf}"
            
            buckets.setdefault(grp, []).append({"name": name, "shift": label})

            if grp == active_group:
                buckets_now.setdefault(grp, []).append({"name": name, "shift": label})

        all_shifts_by_dept.append({"dept": dept_name, "shifts": buckets})

        if dept_name == "Unassigned":
            dept_color = UNASSIGNED_COLOR
        else:
            dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]

        open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None
        card_all = dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full)
        dept_cards_all.append(card_all)

        card_now = dept_card_html(dept_name, dept_color, buckets_now, open_group=active_group)
        dept_cards_now.append(card_now)

        employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
        employees_total_now += sum(len(buckets_now.get(g, [])) for g in GROUP_ORDER)

        depts_count += 1

    os.makedirs("docs", exist_ok=True)
    os.makedirs("docs/now", exist_ok=True)

    try:
        date_label = now.strftime("%-d %B %Y")
    except Exception:
        date_label = now.strftime("%d %B %Y")

    iso_date = now.strftime("%Y-%m-%d")
    sent_time = now.strftime("%H:%M")

    full_url = f"{pages_base}/"
    now_url = f"{pages_base}/now/"

    html_full = page_shell_html(
        date_label=date_label,
        iso_date=iso_date,
        employees_total=employees_total_all,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_all),
        cta_url=now_url,
        sent_time=sent_time,
    )
    html_now = page_shell_html(
        date_label=date_label,
        iso_date=iso_date,
        employees_total=employees_total_now,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_now),
        cta_url=full_url,
        sent_time=sent_time,
    )

    with open("docs/index.html", "w", encoding="utf-8") as f:
        f.write(html_full)

    with open("docs/now/index.html", "w", encoding="utf-8") as f:
        f.write(html_now)

    # Email: send all shifts for today
    subject = f"Duty Roster â€” {now.strftime('%d %B %Y')} â€” {active_group} Active"
    email_html = build_pretty_email_html(active_group, now, all_shifts_by_dept, pages_base)
    send_email(subject, email_html)


if __name__ == "__main__":
    main()